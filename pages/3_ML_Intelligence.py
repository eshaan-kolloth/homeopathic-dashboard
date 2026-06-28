import streamlit as st
if not st.session_state.get("password_correct", False):
    st.stop()

import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from io import BytesIO
import os
import requests
import warnings
warnings.filterwarnings("ignore")

try:
    from sklearn.preprocessing import LabelEncoder, StandardScaler
    from sklearn.model_selection import train_test_split
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.cluster import KMeans
    from sklearn.metrics import accuracy_score
    from sklearn.utils import resample
    from sklearn.linear_model import LinearRegression
    SKLEARN_AVAILABLE = True
except ImportError as _sklearn_err:
    SKLEARN_AVAILABLE = False
    _sklearn_import_error = str(_sklearn_err)

st.set_page_config(
    page_title="ARIA · ML Intelligence",
    page_icon="🔮",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════════════════════════════
# GUARD
# ══════════════════════════════════════════════════════════════════
if "df" not in st.session_state:
    st.warning("No dataset loaded. Please upload data first.")
    if st.button("← Upload Data"):
        st.switch_page("app.py")
    st.stop()

# ══════════════════════════════════════════════════════════════════
# SESSION STATE DEFAULTS
# ══════════════════════════════════════════════════════════════════
if "ml_show_guide" not in st.session_state:
    st.session_state["ml_show_guide"] = True
if "ml_chat_history" not in st.session_state:
    st.session_state["ml_chat_history"] = []
if "ml_active_model" not in st.session_state:
    st.session_state["ml_active_model"] = None   # None = overview, else "risk"/"forecast"/"channel"/"segments"

# ══════════════════════════════════════════════════════════════════
# CSS — Neural Glass theme (kept) + vivid per-model banner system
# ══════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;600&family=Space+Grotesk:wght@400;500;600;700&display=swap');

*,*::before,*::after{box-sizing:border-box;}
:root{
  --bg:#020510;--bg2:#040A1A;--bg3:#071020;
  --surface:rgba(255,255,255,0.032);--surface2:rgba(255,255,255,0.055);
  --border:rgba(255,255,255,0.072);--border2:rgba(255,255,255,0.12);
  --blue:#3B82F6;--blue2:#60A5FA;--cyan:#06B6D4;--purple:#8B5CF6;
  --green:#10B981;--gold:#F59E0B;--red:#EF4444;--orange:#F97316;
  --pink:#EC4899;--lime:#84CC16;
  --txt1:#F1F5F9;--txt2:#94A3B8;--txt3:#475569;
  --radius:16px;--radius-sm:10px;
}
html,body,.stApp{
  background:
    radial-gradient(ellipse 90% 55% at 15% -10%,rgba(139,92,246,0.12) 0%,transparent 60%),
    radial-gradient(ellipse 70% 50% at 85% 105%,rgba(59,130,246,0.09) 0%,transparent 60%),
    radial-gradient(ellipse 50% 40% at 50% 50%,rgba(6,182,212,0.04) 0%,transparent 70%),
    #020510 !important;
  color:var(--txt2);font-family:'Inter',sans-serif;
}
#MainMenu,footer,header{visibility:hidden;}
.block-container{padding:5.5rem 2.4vw 3rem !important;max-width:100% !important;}

/* ══ TOPBAR ══ */
#topbar{
  position:fixed;top:0;left:0;right:0;z-index:300;
  display:flex;align-items:center;justify-content:space-between;
  padding:.75rem 2rem;
  background:rgba(2,5,16,0.88);
  border-bottom:1px solid var(--border);
  backdrop-filter:blur(32px);
}
.tb-brand{display:flex;align-items:center;gap:.7rem;}
.tb-mark{
  width:32px;height:32px;border-radius:9px;
  background:linear-gradient(135deg,#8B5CF6 0%,#3B82F6 55%,#06B6D4 100%);
  display:flex;align-items:center;justify-content:center;
  font-size:.78rem;font-weight:900;color:#fff;
  box-shadow:0 0 20px rgba(139,92,246,.5);
}
.tb-name{font-family:'Space Grotesk',sans-serif;font-size:1rem;font-weight:700;color:var(--txt1);letter-spacing:-.4px;}
.tb-sep{color:var(--border2);margin:0 .3rem;}
.tb-page{font-size:.82rem;color:var(--txt2);font-weight:500;}
.tb-right{display:flex;align-items:center;gap:.55rem;}
.tb-tag{
  font-family:'JetBrains Mono',monospace;font-size:.56rem;
  padding:3px 11px;border-radius:6px;
  background:var(--surface);border:1px solid var(--border);
  color:var(--txt3);letter-spacing:.4px;
}
.tb-live{
  display:flex;align-items:center;gap:5px;
  font-family:'JetBrains Mono',monospace;font-size:.56rem;
  padding:3px 11px;border-radius:6px;
  background:rgba(139,92,246,.1);border:1px solid rgba(139,92,246,.3);
  color:#A78BFA;letter-spacing:.3px;
}
.live-dot{width:5px;height:5px;border-radius:50%;background:#A78BFA;box-shadow:0 0 7px #A78BFA;animation:pulse 2s infinite;}
@keyframes pulse{0%,100%{opacity:1;transform:scale(1);}50%{opacity:.3;transform:scale(.8);}}
@keyframes float{0%,100%{transform:translateY(0px);}50%{transform:translateY(-6px);}}
@keyframes shimmer{0%{background-position:-200% 0;}100%{background-position:200% 0;}}
@keyframes glow-pulse{0%,100%{box-shadow:0 0 20px rgba(139,92,246,.35);}50%{box-shadow:0 0 38px rgba(139,92,246,.6);}}
@keyframes fadeUp{from{opacity:0;transform:translateY(10px);}to{opacity:1;transform:translateY(0);}}

/* ══ SIDEBAR ══ */
section[data-testid="stSidebar"]{
  background:rgba(3,8,22,0.97) !important;
  border-right:1px solid var(--border) !important;
  backdrop-filter:blur(24px) !important;
}
section[data-testid="stSidebar"] *{color:var(--txt2) !important;}
.sb-brand{display:flex;align-items:center;gap:.6rem;padding:.8rem .9rem .7rem;border-bottom:1px solid var(--border);margin-bottom:.8rem;}
.sb-mark{width:32px;height:32px;border-radius:9px;background:linear-gradient(135deg,#8B5CF6,#3B82F6 55%,#06B6D4);display:flex;align-items:center;justify-content:center;font-size:.78rem;font-weight:900;color:#fff;box-shadow:0 0 16px rgba(139,92,246,.4);flex-shrink:0;}
.sb-name{font-family:'Space Grotesk',sans-serif;font-size:.88rem;font-weight:700;color:var(--txt1) !important;}
.sb-ver{font-family:'JetBrains Mono',monospace;font-size:.48rem;color:var(--txt3) !important;}
.sb-section{display:flex;align-items:center;gap:8px;padding:.5rem .9rem .25rem;margin-top:.3rem;}
.sb-section-label{font-family:'JetBrains Mono',monospace;font-size:.5rem;letter-spacing:2.8px;text-transform:uppercase;color:#A78BFA !important;font-weight:600;}
.sb-section-line{flex:1;height:1px;background:linear-gradient(90deg,rgba(139,92,246,.25),transparent);}
.model-card{margin:.2rem .5rem .4rem;background:rgba(255,255,255,.018);border:1px solid rgba(255,255,255,.07);border-radius:12px;padding:.55rem .7rem;transition:all .2s;}
.model-card:hover{background:rgba(139,92,246,.06);border-color:rgba(139,92,246,.2);}
.model-card-top{display:flex;align-items:center;justify-content:space-between;margin-bottom:3px;}
.model-name{font-size:.72rem;font-weight:700;color:var(--txt1) !important;}
.model-desc{font-size:.62rem;color:var(--txt3) !important;line-height:1.4;}
.acc-badge{padding:2px 8px;border-radius:99px;font-family:'JetBrains Mono',monospace;font-size:.52rem;}
.acc-good{background:rgba(16,185,129,.1);color:#10B981;border:1px solid rgba(16,185,129,.2);}
.acc-warn{background:rgba(245,158,11,.1);color:#F59E0B;border:1px solid rgba(245,158,11,.2);}
.acc-low{background:rgba(239,68,68,.1);color:#EF4444;border:1px solid rgba(239,68,68,.2);}

.stButton button,.stDownloadButton button{
  background:linear-gradient(135deg,rgba(139,92,246,.12),rgba(59,130,246,.09)) !important;
  border:1px solid rgba(139,92,246,.25) !important;color:var(--txt1) !important;
  border-radius:11px !important;font-family:'Inter',sans-serif !important;
  font-weight:600 !important;transition:all .22s !important;font-size:.82rem !important;
}
.stButton button:hover{border-color:#A78BFA !important;box-shadow:0 0 20px rgba(139,92,246,.15) !important;transform:translateY(-1px) !important;}

/* ══ HERO BANNER ══ */
.ml-hero{
  border-radius:22px;border:1px solid rgba(139,92,246,.25);
  background:linear-gradient(135deg,rgba(139,92,246,.1) 0%,rgba(59,130,246,.07) 50%,rgba(6,182,212,.06) 100%);
  backdrop-filter:blur(20px);padding:2rem 2.2rem 1.8rem;position:relative;overflow:hidden;margin-bottom:1.4rem;
}
.ml-hero::before{content:'';position:absolute;top:-80px;right:-60px;width:400px;height:400px;border-radius:50%;background:radial-gradient(circle,rgba(139,92,246,.15) 0%,transparent 70%);pointer-events:none;}
.ml-hero-status{display:flex;align-items:center;gap:8px;font-family:'JetBrains Mono',monospace;font-size:.58rem;color:#A78BFA;letter-spacing:2px;text-transform:uppercase;margin-bottom:.8rem;}
.ml-hero-dot{width:6px;height:6px;border-radius:50%;background:#A78BFA;box-shadow:0 0 8px #A78BFA;animation:pulse 2s infinite;}
.ml-hero-title{font-family:'Space Grotesk',sans-serif;font-size:2rem;font-weight:700;color:var(--txt1);margin-bottom:.5rem;letter-spacing:-.6px;line-height:1.15;}
.ml-hero-title span{background:linear-gradient(135deg,#A78BFA,#60A5FA,#06B6D4);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;}
.ml-hero-desc{font-size:.88rem;color:var(--txt3);line-height:1.65;max-width:560px;margin-bottom:1.4rem;}
.ml-hero-chips{display:flex;flex-wrap:wrap;gap:.5rem;}
.ml-chip{display:flex;align-items:center;gap:6px;font-family:'JetBrains Mono',monospace;font-size:.58rem;padding:5px 13px;border-radius:9px;background:rgba(255,255,255,.03);border:1px solid var(--border);color:var(--txt3);}
.ml-chip b{color:var(--txt1);}

/* ══ SECTION ══ */
.sec-block{margin:2.5rem 0 1.3rem;}
.sec-eyebrow{display:flex;align-items:center;gap:10px;font-family:'JetBrains Mono',monospace;font-size:.6rem;color:#A78BFA;letter-spacing:2.8px;text-transform:uppercase;margin-bottom:.55rem;}
.sec-eyebrow .line{flex:1;height:1px;background:linear-gradient(90deg,rgba(139,92,246,.3),transparent);}
.sec-title{font-family:'Space Grotesk',sans-serif;font-size:1.4rem;font-weight:700;color:var(--txt1);line-height:1.2;margin-bottom:.4rem;letter-spacing:-.5px;}
.sec-subtitle{font-size:.88rem;color:var(--txt3);line-height:1.65;max-width:720px;padding-bottom:.6rem;border-bottom:1px solid rgba(255,255,255,.04);margin-bottom:.2rem;}

/* ══ GLASS CARD ══ */
.gc{border-radius:18px;border:1px solid var(--border);background:var(--surface);backdrop-filter:blur(16px);padding:1.3rem 1.4rem;height:100%;}
.gc-title{font-family:'JetBrains Mono',monospace;font-size:.56rem;color:var(--txt3);text-transform:uppercase;letter-spacing:1.6px;margin-bottom:1rem;display:flex;align-items:center;gap:7px;}
.gc-dot{width:6px;height:6px;border-radius:50%;flex-shrink:0;}

/* ══ KPI ══ */
.kpi{border-radius:18px;border:1px solid var(--border);background:var(--surface);padding:1.2rem 1.3rem 1.1rem;position:relative;overflow:hidden;transition:transform .25s ease,border-color .25s ease;}
.kpi::before{content:'';position:absolute;top:0;left:0;right:0;height:2.5px;border-radius:18px 18px 0 0;}
.kpi.purple::before{background:linear-gradient(90deg,#8B5CF6,#3B82F6);}
.kpi.blue::before{background:linear-gradient(90deg,#3B82F6,#06B6D4);}
.kpi.green::before{background:linear-gradient(90deg,#10B981,#06B6D4);}
.kpi.gold::before{background:linear-gradient(90deg,#F59E0B,#F97316);}
.kpi.red::before{background:linear-gradient(90deg,#EF4444,#F97316);}
.kpi.pink::before{background:linear-gradient(90deg,#EC4899,#8B5CF6);}
.kpi:hover{transform:translateY(-4px);border-color:var(--border2);}
.kpi-label{font-family:'JetBrains Mono',monospace;font-size:.52rem;color:var(--txt3);text-transform:uppercase;letter-spacing:1.4px;margin-bottom:.35rem;}
.kpi-value{font-family:'Space Grotesk',sans-serif;font-size:1.65rem;font-weight:700;line-height:1;margin-bottom:.3rem;letter-spacing:-.5px;}
.kpi-value.purple{color:#A78BFA;}.kpi-value.blue{color:#60A5FA;}.kpi-value.green{color:#10B981;}
.kpi-value.gold{color:#F59E0B;}.kpi-value.red{color:#EF4444;}.kpi-value.pink{color:#EC4899;}
.kpi-sub{font-size:.7rem;color:var(--txt3);line-height:1.4;}

/* ══ RISK TABLE ══ */
.risk-row{display:flex;align-items:center;gap:10px;padding:.5rem .7rem;border-radius:10px;margin-bottom:.3rem;background:rgba(255,255,255,.018);border:1px solid rgba(255,255,255,.05);transition:all .15s;}
.risk-row:hover{background:rgba(255,255,255,.03);}
.risk-badge{font-family:'JetBrains Mono',monospace;font-size:.52rem;padding:3px 10px;border-radius:99px;flex-shrink:0;font-weight:700;}
.rb-paid{background:rgba(16,185,129,.1);color:#10B981;border:1px solid rgba(16,185,129,.22);}
.rb-risk{background:rgba(239,68,68,.1);color:#EF4444;border:1px solid rgba(239,68,68,.22);}
.rb-partial{background:rgba(245,158,11,.1);color:#F59E0B;border:1px solid rgba(245,158,11,.22);}
.risk-party{font-size:.8rem;color:var(--txt1);font-weight:600;flex:1;min-width:0;}
.risk-amt{font-family:'JetBrains Mono',monospace;font-size:.75rem;color:var(--txt2);flex-shrink:0;}
.risk-due{font-family:'JetBrains Mono',monospace;font-size:.68rem;color:#EF4444;flex-shrink:0;}

/* ══ SEGMENT CARD ══ */
.seg-card{border-radius:16px;border:1px solid var(--border);background:var(--surface);padding:1.1rem 1.2rem;margin-bottom:.6rem;transition:all .2s;}
.seg-card:hover{transform:translateY(-2px);border-color:var(--border2);}
.seg-card-top{display:flex;align-items:center;justify-content:space-between;margin-bottom:.6rem;}
.seg-name{font-size:.9rem;font-weight:700;color:var(--txt1);}
.seg-label{font-family:'JetBrains Mono',monospace;font-size:.52rem;padding:3px 10px;border-radius:99px;font-weight:700;}
.sl-vip{background:rgba(245,158,11,.1);color:#F59E0B;border:1px solid rgba(245,158,11,.2);}
.sl-regular{background:rgba(59,130,246,.1);color:#60A5FA;border:1px solid rgba(59,130,246,.2);}
.sl-risk{background:rgba(239,68,68,.1);color:#EF4444;border:1px solid rgba(239,68,68,.2);}
.seg-stats{display:grid;grid-template-columns:repeat(3,1fr);gap:.4rem;}
.seg-stat{text-align:center;}
.seg-stat-val{font-family:'JetBrains Mono',monospace;font-size:.78rem;font-weight:700;color:var(--txt1);}
.seg-stat-lbl{font-size:.6rem;color:var(--txt3);margin-top:1px;}
.seg-action{margin-top:.7rem;padding:.5rem .7rem;border-radius:9px;font-size:.75rem;color:var(--txt2);background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.05);}

/* ══ CHANNEL TREND ══ */
.ch-trend-row{display:flex;align-items:center;gap:.7rem;padding:.45rem 0;border-bottom:1px solid rgba(255,255,255,.03);}
.ch-trend-row:last-child{border-bottom:none;}
.ch-icon{font-size:1rem;flex-shrink:0;}
.ch-name{font-size:.8rem;color:var(--txt1);font-weight:600;flex:1;}

/* ══ INSIGHT BOX ══ */
.insight-box{border-radius:13px;padding:.9rem 1.1rem;margin-top:.8rem;font-size:.8rem;line-height:1.7;color:var(--txt2);}
.insight-box.purple{background:rgba(139,92,246,.06);border:1px solid rgba(139,92,246,.15);}
.insight-box.blue{background:rgba(59,130,246,.06);border:1px solid rgba(59,130,246,.15);}
.insight-box.green{background:rgba(16,185,129,.06);border:1px solid rgba(16,185,129,.15);}
.insight-box.gold{background:rgba(245,158,11,.06);border:1px solid rgba(245,158,11,.15);}
.insight-box.red{background:rgba(239,68,68,.06);border:1px solid rgba(239,68,68,.15);}

/* ══ FOOTER ══ */
.aria-footer{text-align:center;margin-top:3rem;padding-top:1.5rem;border-top:1px solid var(--border);font-family:'JetBrains Mono',monospace;font-size:.58rem;color:var(--txt3);letter-spacing:1.2px;}
::-webkit-scrollbar{width:4px;height:4px;}
::-webkit-scrollbar-track{background:transparent;}
::-webkit-scrollbar-thumb{background:rgba(139,92,246,.3);border-radius:99px;}

/* ════════════════════════════════════════════════════════════════
   WELCOME GUIDE (kept as-is — user liked this)
   ════════════════════════════════════════════════════════════════ */
.guide-wrap{
  border-radius:24px;border:1.5px solid rgba(167,139,250,.4);
  background:linear-gradient(135deg,rgba(139,92,246,.16) 0%,rgba(59,130,246,.10) 45%,rgba(6,182,212,.08) 100%);
  backdrop-filter:blur(24px);padding:1.9rem 2.1rem 1.7rem;position:relative;overflow:hidden;margin-bottom:1.6rem;
  animation:glow-pulse 4s ease-in-out infinite;
}
.guide-wrap::before{content:'✨';position:absolute;top:-30px;right:20px;font-size:8rem;opacity:.06;transform:rotate(15deg);pointer-events:none;}
.guide-badge{display:inline-flex;align-items:center;gap:7px;font-family:'JetBrains Mono',monospace;font-size:.58rem;letter-spacing:2px;text-transform:uppercase;padding:5px 14px;border-radius:99px;margin-bottom:.9rem;background:linear-gradient(90deg,rgba(167,139,250,.25),rgba(96,165,250,.25),rgba(167,139,250,.25));background-size:200% 100%;animation:shimmer 3s linear infinite;border:1px solid rgba(167,139,250,.45);color:#fff;font-weight:700;}
.guide-title{font-family:'Space Grotesk',sans-serif;font-size:1.55rem;font-weight:800;color:var(--txt1);margin-bottom:.5rem;letter-spacing:-.5px;line-height:1.2;}
.guide-title span{background:linear-gradient(135deg,#F59E0B,#EC4899,#A78BFA);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;}
.guide-desc{font-size:.88rem;color:#CBD5E1;line-height:1.7;max-width:740px;margin-bottom:1.3rem;}
.guide-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:.85rem;margin-bottom:1rem;}
.guide-step{border-radius:16px;padding:1rem 1.05rem;position:relative;background:rgba(255,255,255,.045);border:1px solid rgba(255,255,255,.09);transition:all .25s;}
.guide-step:hover{transform:translateY(-4px);background:rgba(255,255,255,.07);}
.guide-step-num{position:absolute;top:-10px;left:14px;width:26px;height:26px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-family:'Space Grotesk',sans-serif;font-weight:800;font-size:.72rem;color:#fff;box-shadow:0 4px 12px rgba(0,0,0,.3);}
.guide-step-icon{font-size:1.6rem;margin:.4rem 0 .55rem;display:block;}
.guide-step-title{font-size:.84rem;font-weight:700;color:var(--txt1);margin-bottom:.3rem;}
.guide-step-desc{font-size:.72rem;color:#94A3B8;line-height:1.55;}
.guide-footer-row{display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:.7rem;margin-top:.5rem;}
.guide-hint{font-size:.74rem;color:#A78BFA;display:flex;align-items:center;gap:6px;}

/* ════════════════════════════════════════════════════════════════
   MODEL LAUNCHER BUTTONS — the 4 big interactive cards
   ════════════════════════════════════════════════════════════════ */
.launcher-card{
  border-radius:20px;padding:1.4rem 1.3rem 1.2rem;height:100%;cursor:pointer;
  border:1.5px solid var(--border2);position:relative;overflow:hidden;
  transition:all .25s ease;
}
.launcher-card.lc-risk{background:linear-gradient(145deg,rgba(239,68,68,.13),rgba(236,72,153,.07));border-color:rgba(239,68,68,.32);}
.launcher-card.lc-forecast{background:linear-gradient(145deg,rgba(139,92,246,.13),rgba(59,130,246,.07));border-color:rgba(139,92,246,.32);}
.launcher-card.lc-channel{background:linear-gradient(145deg,rgba(6,182,212,.13),rgba(59,130,246,.07));border-color:rgba(6,182,212,.32);}
.launcher-card.lc-segment{background:linear-gradient(145deg,rgba(245,158,11,.13),rgba(236,72,153,.07));border-color:rgba(245,158,11,.32);}
.launcher-card.active-glow{box-shadow:0 0 0 2px #fff inset,0 8px 30px rgba(0,0,0,.4);}
.launcher-icon{font-size:2.1rem;margin-bottom:.5rem;display:block;}
.launcher-title{font-family:'Space Grotesk',sans-serif;font-size:1.02rem;font-weight:800;color:var(--txt1);margin-bottom:.3rem;}
.launcher-desc{font-size:.74rem;color:#CBD5E1;line-height:1.55;margin-bottom:.7rem;}
.launcher-stat{font-family:'JetBrains Mono',monospace;font-size:.62rem;padding:3px 10px;border-radius:99px;display:inline-block;background:rgba(255,255,255,.1);color:#fff;font-weight:700;}
.launcher-cta{margin-top:.7rem;font-family:'JetBrains Mono',monospace;font-size:.6rem;letter-spacing:1px;text-transform:uppercase;color:#fff;opacity:.85;display:flex;align-items:center;gap:5px;}

/* ════════════════════════════════════════════════════════════════
   PER-MODEL DETAIL BANNERS — unique color per model
   ════════════════════════════════════════════════════════════════ */
.model-banner{
  border-radius:22px;padding:1.8rem 2rem;position:relative;overflow:hidden;margin-bottom:1.6rem;
  backdrop-filter:blur(20px);animation:fadeUp .4s ease;
}
.model-banner::before{content:'';position:absolute;bottom:-100px;left:-60px;width:340px;height:340px;border-radius:50%;pointer-events:none;}
.mb-risk{border:1px solid rgba(239,68,68,.3);background:linear-gradient(135deg,rgba(239,68,68,.1) 0%,rgba(236,72,153,.07) 50%,rgba(139,92,246,.06) 100%);}
.mb-risk::before{background:radial-gradient(circle,rgba(239,68,68,.16) 0%,transparent 70%);}
.mb-forecast{border:1px solid rgba(139,92,246,.3);background:linear-gradient(135deg,rgba(139,92,246,.1) 0%,rgba(59,130,246,.07) 50%,rgba(6,182,212,.06) 100%);}
.mb-forecast::before{background:radial-gradient(circle,rgba(139,92,246,.16) 0%,transparent 70%);}
.mb-channel{border:1px solid rgba(6,182,212,.3);background:linear-gradient(135deg,rgba(6,182,212,.1) 0%,rgba(59,130,246,.07) 50%,rgba(16,185,129,.06) 100%);}
.mb-channel::before{background:radial-gradient(circle,rgba(6,182,212,.16) 0%,transparent 70%);}
.mb-segment{border:1px solid rgba(245,158,11,.3);background:linear-gradient(135deg,rgba(245,158,11,.1) 0%,rgba(236,72,153,.07) 50%,rgba(139,92,246,.06) 100%);}
.mb-segment::before{background:radial-gradient(circle,rgba(245,158,11,.16) 0%,transparent 70%);}
.mb-outlook{border:1px solid rgba(16,185,129,.28);background:linear-gradient(135deg,rgba(16,185,129,.09) 0%,rgba(6,182,212,.07) 50%,rgba(59,130,246,.06) 100%);}
.mb-outlook::before{background:radial-gradient(circle,rgba(16,185,129,.14) 0%,transparent 70%);}

.mb-head{display:flex;align-items:center;gap:10px;margin-bottom:.5rem;position:relative;z-index:1;}
.mb-eyebrow{font-family:'JetBrains Mono',monospace;font-size:.58rem;letter-spacing:2.4px;text-transform:uppercase;}
.mb-title{font-family:'Space Grotesk',sans-serif;font-size:1.45rem;font-weight:800;color:var(--txt1);margin-bottom:.8rem;position:relative;z-index:1;}
.mb-narrative{font-size:.92rem;color:#E2E8F0;line-height:1.85;position:relative;z-index:1;max-width:920px;}
.mb-narrative b{font-weight:700;}
.mb-pills{display:flex;flex-wrap:wrap;gap:.6rem;margin-top:1.1rem;position:relative;z-index:1;}
.mb-pill{display:flex;align-items:center;gap:7px;padding:7px 15px;border-radius:99px;background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.12);font-size:.76rem;color:var(--txt1);font-weight:600;}

/* ════════════════════════════════════════════════════════════════
   STEP-BY-STEP "HOW ARIA THINKS" CARDS
   ════════════════════════════════════════════════════════════════ */
.step-card{border-radius:16px;padding:1rem 1.15rem;margin-bottom:.6rem;background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.08);display:flex;gap:12px;align-items:flex-start;}
.step-num{flex-shrink:0;width:30px;height:30px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-family:'Space Grotesk',sans-serif;font-weight:800;font-size:.78rem;color:#fff;}
.step-body{flex:1;}
.step-title{font-size:.84rem;font-weight:700;color:var(--txt1);margin-bottom:.25rem;}
.step-desc{font-size:.76rem;color:#94A3B8;line-height:1.6;}

/* ════════════════════════════════════════════════════════════════
   GLOSSARY MINI CARDS
   ════════════════════════════════════════════════════════════════ */
.gloss-row{display:flex;gap:10px;margin-bottom:.7rem;padding-bottom:.6rem;border-bottom:1px solid rgba(255,255,255,.04);}
.gloss-icon{font-size:1.1rem;flex-shrink:0;}
.gloss-term{font-size:.78rem;font-weight:700;margin-bottom:.2rem;}
.gloss-def{font-size:.71rem;color:var(--txt3);line-height:1.5;}

/* ════════════════════════════════════════════════════════════════
   AI CHAT DOCK — creative floating-card style, sticky-feel column
   ════════════════════════════════════════════════════════════════ */
.chat-wrap{
  border-radius:22px;border:1.5px solid rgba(139,92,246,.35);
  background:linear-gradient(160deg,rgba(139,92,246,.1),rgba(6,182,212,.05) 60%,rgba(2,5,16,.4));
  padding:1.3rem 1.35rem;position:relative;overflow:hidden;
  box-shadow:0 10px 40px rgba(139,92,246,.12);
}
.chat-wrap::before{content:'🔮';position:absolute;top:-20px;right:-10px;font-size:6rem;opacity:.05;pointer-events:none;}
.chat-head{display:flex;align-items:center;gap:9px;margin-bottom:.25rem;position:relative;z-index:1;}
.chat-title{font-family:'Space Grotesk',sans-serif;font-size:1rem;font-weight:800;color:var(--txt1);}
.chat-badge{font-family:'JetBrains Mono',monospace;font-size:.5rem;padding:3px 9px;border-radius:99px;background:rgba(139,92,246,.18);border:1px solid rgba(139,92,246,.4);color:#A78BFA;margin-left:auto;}
.chat-desc{font-size:.72rem;color:var(--txt3);margin-bottom:.85rem;line-height:1.5;position:relative;z-index:1;}
.chat-scope-tag{
  display:inline-flex;align-items:center;gap:5px;font-family:'JetBrains Mono',monospace;font-size:.55rem;
  padding:3px 10px;border-radius:99px;background:rgba(16,185,129,.1);border:1px solid rgba(16,185,129,.3);
  color:#34D399;margin-bottom:.7rem;position:relative;z-index:1;
}
.chat-bubble{border-radius:13px;padding:.65rem .85rem;margin-bottom:.5rem;font-size:.77rem;line-height:1.55;position:relative;z-index:1;}
.chat-bubble.user{background:rgba(59,130,246,.12);border:1px solid rgba(59,130,246,.25);color:#DBEAFE;margin-left:1rem;}
.chat-bubble.ai{background:rgba(139,92,246,.1);border:1px solid rgba(139,92,246,.22);color:#E9D5FF;margin-right:.3rem;}
.chat-bubble-label{font-family:'JetBrains Mono',monospace;font-size:.48rem;text-transform:uppercase;letter-spacing:1px;opacity:.7;margin-bottom:3px;display:block;}
.chat-empty{font-size:.74rem;color:var(--txt3);padding:.5rem 0 .7rem;position:relative;z-index:1;line-height:1.6;}

html{scroll-behavior:smooth;}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════
# LOAD DATA
# ══════════════════════════════════════════════════════════════════
df = st.session_state["df"]
filename = st.session_state.get("filename", "dataset.xlsx")

tx_df = pd.DataFrame()
if "uploaded_file_bytes" in st.session_state:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(st.session_state["uploaded_file_bytes"]), data_only=True)
        if "Transactions" in wb.sheetnames:
            ws = wb["Transactions"]
            rows = list(ws.iter_rows(values_only=True))
            hdrs = rows[0]
            tx_df = pd.DataFrame([dict(zip(hdrs, r)) for r in rows[1:] if any(v is not None for v in r)])
    except Exception:
        pass

# ══════════════════════════════════════════════════════════════════
# ML ENGINE (model logic unchanged)
# ══════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False, ttl=600)
def run_ml_models(cache_key: str):
    results = {}
    if not SKLEARN_AVAILABLE:
        results["sklearn_error"] = _sklearn_import_error
        return results
    if tx_df.empty or "Total Amount" not in tx_df.columns:
        return results

    df_ml = tx_df.copy()
    try:
        df_ml["Date"] = pd.to_datetime(df_ml["Date"], dayfirst=True, errors="coerce")
        df_ml["Follow up date"] = pd.to_datetime(df_ml["Follow up date"], dayfirst=True, errors="coerce")
    except Exception:
        pass

    df_ml["DayOfWeek"]    = df_ml["Date"].dt.dayofweek.fillna(0).astype(int)
    df_ml["DayOfMonth"]   = df_ml["Date"].dt.day.fillna(1).astype(int)
    df_ml["FollowUpDays"] = ((df_ml["Follow up date"] - df_ml["Date"]).dt.days).fillna(3).astype(int)
    df_ml["IsPaid"]       = (df_ml["Balance Due"] == 0).astype(int)
    df_ml["PaymentRatio"] = (df_ml["Received/Paid Amount"] / df_ml["Total Amount"].replace(0, 1)).fillna(0)

    le_pay, le_ch, le_sm, le_par = LabelEncoder(), LabelEncoder(), LabelEncoder(), LabelEncoder()
    df_ml["PaymentType_enc"]  = le_pay.fit_transform(df_ml["Payment Type"].astype(str))
    df_ml["SalesChannel_enc"] = le_ch.fit_transform(df_ml["Sales Channel"].astype(str))
    df_ml["Salesman_enc"]     = le_sm.fit_transform(df_ml["Salesman Name"].astype(str))
    df_ml["PartyName_enc"]    = le_par.fit_transform(df_ml["Party Name"].astype(str))

    base_feats = ["Total Amount","PaymentType_enc","SalesChannel_enc",
                  "Salesman_enc","PartyName_enc","DayOfWeek","DayOfMonth","FollowUpDays"]

    # MODEL 1: Payment Risk
    try:
        X1 = df_ml[base_feats].fillna(0)
        y1 = df_ml["IsPaid"]
        df_maj, df_min = df_ml[y1 == 0], df_ml[y1 == 1]
        if len(df_min) > 0 and len(df_maj) > 0:
            df_min_up = resample(df_min, replace=True, n_samples=len(df_maj), random_state=42)
            df_bal = pd.concat([df_maj, df_min_up])
            X1b, y1b = df_bal[base_feats].fillna(0), df_bal["IsPaid"]
        else:
            X1b, y1b = X1, y1
        X1_tr, X1_te, y1_tr, y1_te = train_test_split(X1b, y1b, test_size=0.2, random_state=42)
        m1 = RandomForestClassifier(n_estimators=100, random_state=42, class_weight="balanced")
        m1.fit(X1_tr, y1_tr)
        acc1 = accuracy_score(y1_te, m1.predict(X1_te))
        proba = m1.predict_proba(X1.fillna(0))
        paid_idx = list(m1.classes_).index(1) if 1 in m1.classes_ else 0
        df_ml["RiskScore"] = proba[:, 1 - paid_idx] if len(m1.classes_) > 1 else 0.5
        df_ml["PayPrediction"] = m1.predict(X1.fillna(0))
        results["model1"] = {
            "accuracy": acc1,
            "df": df_ml[["Party Name","Salesman Name","Sales Channel","Total Amount",
                         "Received/Paid Amount","Balance Due","Payment Type","IsPaid",
                         "RiskScore","PayPrediction","Follow up date"]].copy(),
            "feature_importance": dict(zip(base_feats, m1.feature_importances_)),
        }
    except Exception as e:
        results["model1_error"] = str(e)

    # MODEL 2: Revenue Forecast
    try:
        EXCLUDE_KW = ["total","net sales","return","month","year","date","agri","precription",
                      "prescription","coupon","discount","gst","transporation","bill amount"]
        person_mask = df.iloc[:, 0].astype(str).apply(
            lambda x: not any(kw in x.lower() for kw in EXCLUDE_KW)
                      and x.strip().lower() not in ("0","","nan","daily report")
        )
        person_df_m2 = df[person_mask].copy()
        numeric_m2 = person_df_m2.iloc[:, 1:].apply(pd.to_numeric, errors="coerce").fillna(0)
        daily_vals = numeric_m2.sum(axis=0).values

        if len(daily_vals) >= 5:
            X_time = np.arange(len(daily_vals)).reshape(-1, 1)
            y_time = daily_vals
            reg = LinearRegression()
            reg.fit(X_time, y_time)
            n_future = 7
            X_future = np.arange(len(daily_vals), len(daily_vals) + n_future).reshape(-1, 1)
            forecast = np.maximum(reg.predict(X_future), 0)
            conf_lo, conf_hi = forecast * 0.85, forecast * 1.15
            residuals = y_time - reg.predict(X_time)
            rmse = np.sqrt(np.mean(residuals**2))
            results["model2"] = {
                "historical": daily_vals.tolist(), "forecast": forecast.tolist(),
                "conf_lo": conf_lo.tolist(), "conf_hi": conf_hi.tolist(),
                "rmse": rmse, "trend_slope": float(reg.coef_[0]),
            }
    except Exception as e:
        results["model2_error"] = str(e)

    # MODEL 3: Channel Intelligence
    try:
        ch_stats = df_ml.groupby("Sales Channel").agg(
            Revenue=("Total Amount", "sum"), Transactions=("Invoice No", "count"),
            AvgOrder=("Total Amount", "mean"), CollectionRate=("IsPaid", "mean"),
            AvgDayOfMonth=("DayOfMonth", "mean"),
        ).reset_index()
        first_h = df_ml[df_ml["DayOfMonth"] <= 15].groupby("Sales Channel")["Total Amount"].sum()
        second_h = df_ml[df_ml["DayOfMonth"] > 15].groupby("Sales Channel")["Total Amount"].sum()
        ch_stats["FirstHalf"] = ch_stats["Sales Channel"].map(first_h).fillna(0)
        ch_stats["SecondHalf"] = ch_stats["Sales Channel"].map(second_h).fillna(0)
        ch_stats["Trend"] = ch_stats["SecondHalf"] - ch_stats["FirstHalf"]
        ch_stats["TrendPct"] = (ch_stats["Trend"] / ch_stats["FirstHalf"].replace(0, 1) * 100)
        results["model3"] = {"df": ch_stats}
    except Exception as e:
        results["model3_error"] = str(e)

    # MODEL 4: Customer Segmentation
    try:
        cust = df_ml.groupby("Party Name").agg(
            Total_Transactions=("Invoice No", "count"), Total_Revenue=("Total Amount", "sum"),
            Avg_Order_Value=("Total Amount", "mean"), Total_Paid=("Received/Paid Amount", "sum"),
            Total_Balance=("Balance Due", "sum"), Payment_Rate=("IsPaid", "mean"),
            Unique_Channels=("Sales Channel", "nunique"),
        ).reset_index()
        scaler = StandardScaler()
        feat_cols = ["Total_Transactions","Total_Revenue","Avg_Order_Value","Payment_Rate","Total_Balance","Unique_Channels"]
        X4 = scaler.fit_transform(cust[feat_cols].fillna(0))
        n_clusters = min(3, len(cust))
        km = KMeans(n_clusters=n_clusters, random_state=42, n_init=10)
        cust["Cluster"] = km.fit_predict(X4)
        cluster_rev = cust.groupby("Cluster")["Total_Revenue"].mean().sort_values(ascending=False)
        rank_map = {c: i for i, c in enumerate(cluster_rev.index)}
        cust["Segment_Rank"] = cust["Cluster"].map(rank_map)
        SEG_LABELS = {0: "VIP", 1: "Regular", 2: "At Risk"}
        cust["Segment"] = cust["Segment_Rank"].map(SEG_LABELS).fillna("Regular")
        results["model4"] = {"df": cust}
    except Exception as e:
        results["model4_error"] = str(e)

    return results


cache_key = f"{filename}_{len(tx_df)}_{df.shape}"
with st.spinner("🔮 Running ML models…"):
    ml_results = run_ml_models(cache_key)

if not SKLEARN_AVAILABLE:
    st.error(
        "⚠️ **scikit-learn is not installed.** ML models cannot run.\n\n"
        "Add `scikit-learn` to your `requirements.txt` and redeploy the app.\n\n"
        f"Details: `{_sklearn_import_error}`"
    )
    st.stop()

# ══════════════════════════════════════════════════════════════════
# CHART THEME
# ══════════════════════════════════════════════════════════════════
PLOT_BASE = dict(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                  font=dict(family="Inter, sans-serif", color="#94A3B8", size=11))
PAL = ["#8B5CF6","#3B82F6","#06B6D4","#10B981","#F59E0B","#EF4444","#EC4899","#F97316"]

def style_fig(fig, height=300, margin=None, **kw):
    fig.update_layout(**PLOT_BASE, height=height, **kw)
    fig.update_layout(margin=margin or dict(l=8, r=8, t=32, b=8))
    return fig


# ══════════════════════════════════════════════════════════════════
# AI HELPERS — Groq backend, shared by outlook + per-model banners + chat
# ══════════════════════════════════════════════════════════════════
def _groq_chat(messages, max_tokens=400, temperature=0.4):
    api_key = st.session_state.get("groq_api_key") or os.environ.get("GROQ_API_KEY", "")
    if not api_key:
        return None
    try:
        resp = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={"model": "llama-3.3-70b-versatile", "messages": messages,
                  "temperature": temperature, "max_tokens": max_tokens},
            timeout=25,
        )
        if resp.status_code == 200:
            return resp.json()["choices"][0]["message"]["content"].strip()
        return None
    except Exception:
        return None


def build_ml_context_text(ml_results: dict) -> str:
    lines = ["=== ARIA ML INTELLIGENCE SNAPSHOT ==="]
    if "model1" in ml_results:
        m1, m1_df = ml_results["model1"], ml_results["model1"]["df"]
        at_risk = m1_df[m1_df["RiskScore"] > 0.5]
        lines.append(f"\n[PAYMENT RISK MODEL] Accuracy: {m1['accuracy']*100:.1f}%")
        lines.append(f"Total transactions: {len(m1_df)} | At-risk: {len(at_risk)}, ₹{at_risk['Balance Due'].sum():,.0f} outstanding")
        for _, r in at_risk.sort_values("RiskScore", ascending=False).head(5).iterrows():
            lines.append(f"  - {r['Party Name']} | risk {r['RiskScore']*100:.0f}% | due ₹{r['Balance Due']:,.0f}")
    if "model2" in ml_results:
        m2 = ml_results["model2"]
        lines.append(f"\n[REVENUE FORECAST] Trend slope: ₹{m2['trend_slope']:+,.0f}/day")
        lines.append(f"Historical total: ₹{sum(m2['historical']):,.0f} over {len(m2['historical'])} days | 7-day forecast: ₹{sum(m2['forecast']):,.0f}")
    if "model3" in ml_results:
        ch_df = ml_results["model3"]["df"]
        lines.append(f"\n[CHANNEL INTELLIGENCE] {len(ch_df)} channels")
        for _, r in ch_df.sort_values("Revenue", ascending=False).iterrows():
            lines.append(f"  - {r['Sales Channel']}: ₹{r['Revenue']:,.0f}, {r['TrendPct']:+.1f}% trend, {r['CollectionRate']*100:.0f}% collected")
    if "model4" in ml_results:
        cust_df = ml_results["model4"]["df"]
        lines.append(f"\n[CUSTOMER SEGMENTATION] {len(cust_df)} customers")
        for seg in ["VIP", "Regular", "At Risk"]:
            seg_rows = cust_df[cust_df["Segment"] == seg]
            if not seg_rows.empty:
                lines.append(f"  - {seg}: {len(seg_rows)} customers, ₹{seg_rows['Total_Revenue'].sum():,.0f}, names: {', '.join(seg_rows['Party Name'].tolist())}")
    return "\n".join(lines)


@st.cache_data(show_spinner=False, ttl=600)
def generate_business_outlook(cache_key: str, context_text: str):
    messages = [
        {"role": "system", "content": (
            "You are ARIA, a friendly business analyst speaking to a non-technical business owner. "
            "Given ML model results, write ONE short paragraph (4-6 sentences) in plain English — "
            "no jargon, no model names. Cover: overall money outlook for next week, who/what needs "
            "attention, and one clear recommended action. Be specific with rupee numbers and names."
        )},
        {"role": "user", "content": context_text[:6000]},
    ]
    return _groq_chat(messages, max_tokens=320, temperature=0.5)


@st.cache_data(show_spinner=False, ttl=600)
def generate_model_narrative(cache_key: str, model_label: str, context_text: str):
    """Per-model plain-English narrative for that model's banner."""
    messages = [
        {"role": "system", "content": (
            f"You are ARIA, a friendly business analyst. Write ONE short paragraph (3-5 sentences) "
            f"in plain English, no jargon, explaining what the '{model_label}' model found in this "
            f"data and what the business owner should do about it. Be specific with numbers and names. "
            f"Sound like a smart colleague giving a quick briefing."
        )},
        {"role": "user", "content": context_text[:6000]},
    ]
    return _groq_chat(messages, max_tokens=260, temperature=0.5)


def generate_chat_reply(user_question: str, context_text: str, history: list, scope_label: str):
    history_msgs = []
    for turn in history[-6:]:
        history_msgs.append({"role": "user", "content": turn["q"]})
        history_msgs.append({"role": "assistant", "content": turn["a"]})
    messages = [
        {"role": "system", "content": (
            "You are ARIA, an AI business copilot embedded inside a sales ML analytics page. "
            f"The user is currently viewing: {scope_label}. "
            "You have access to real ML model outputs (payment risk, revenue forecast, channel "
            "trends, customer segments). Only answer questions about future business outlook, "
            "predictions, risk, forecasts, channels, or customers — politely redirect anything else "
            "back to these topics. Answer in plain, friendly English for someone with NO data science "
            "background. Be concise (2-5 sentences), specific with numbers, and end with a practical "
            f"next step if relevant.\n\nCURRENT DATA SNAPSHOT:\n{context_text[:5000]}"
        )},
    ] + history_msgs + [{"role": "user", "content": user_question}]
    reply = _groq_chat(messages, max_tokens=380, temperature=0.45)
    return reply or "I couldn't reach the AI service just now (no GROQ API key set). The model results below are still fully accurate — just browse the cards and charts for now."


# ══════════════════════════════════════════════════════════════════
# TOPBAR
# ══════════════════════════════════════════════════════════════════
st.markdown(f"""
<div id="topbar">
  <div class="tb-brand">
    <div class="tb-mark">🔮</div>
    <div class="tb-name">ARIA</div>
    <div class="tb-sep">·</div>
    <div class="tb-page">ML Intelligence Engine</div>
  </div>
  <div class="tb-right">
    <div class="tb-tag">{filename}</div>
    <div class="tb-tag">{len(tx_df)} transactions</div>
    <div class="tb-live"><div class="live-dot"></div>4 Models Active</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
    <div class="sb-brand">
      <div class="sb-mark">🔮</div>
      <div><div class="sb-name">ML Intelligence</div><div class="sb-ver">v3.0 · NEURAL GLASS</div></div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="sb-section"><div class="sb-section-label">Need Help?</div><div class="sb-section-line"></div></div>', unsafe_allow_html=True)
    if st.button("✨ Open Welcome Guide", use_container_width=True, key="open_guide_btn"):
        st.session_state["ml_show_guide"] = True
        st.rerun()

    st.markdown('<div class="sb-section" style="margin-top:.6rem;"><div class="sb-section-label">Dataset</div><div class="sb-section-line"></div></div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="model-card">
      <div class="model-card-top"><span class="model-name">📄 {filename[:18]}…</span></div>
      <div class="model-desc">{len(tx_df)} transactions · {len(tx_df.columns) if not tx_df.empty else 0} features</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="sb-section" style="margin-top:.6rem;"><div class="sb-section-label">Navigate</div><div class="sb-section-line"></div></div>', unsafe_allow_html=True)
    if st.button("📊 Dashboard", use_container_width=True):
        st.switch_page("pages/1_Dashboard.py")
    if st.button("🧠 AI Copilot (Full Page)", use_container_width=True):
        st.switch_page("pages/2_AI_Chat.py")
    if st.button("← New Upload", use_container_width=True):
        st.switch_page("app.py")

# ══════════════════════════════════════════════════════════════════
# SHARED CONTEXT
# ══════════════════════════════════════════════════════════════════
ml_context_text = build_ml_context_text(ml_results)

n_at_risk, at_risk_amount = 0, 0.0
if "model1" in ml_results:
    risky_global = ml_results["model1"]["df"]
    risky_global = risky_global[risky_global["RiskScore"] > 0.5]
    n_at_risk = int(len(risky_global))
    at_risk_amount = float(risky_global["Balance Due"].sum())

forecast_7d = sum(ml_results["model2"]["forecast"]) if "model2" in ml_results else 0.0
n_declining_channels = int((ml_results["model3"]["df"]["TrendPct"] < 0).sum()) if "model3" in ml_results else 0

n_vip = n_risk_cust = 0
if "model4" in ml_results:
    seg_counts = ml_results["model4"]["df"]["Segment"].value_counts()
    n_vip = int(seg_counts.get("VIP", 0))
    n_risk_cust = int(seg_counts.get("At Risk", 0))

m1_acc = ml_results.get("model1", {}).get("accuracy", 0)

# ══════════════════════════════════════════════════════════════════
# ① WELCOME GUIDE
# ══════════════════════════════════════════════════════════════════
if st.session_state.get("ml_show_guide", True):
    st.markdown(f"""
    <div class="guide-wrap">
      <div class="guide-badge">✨ Welcome · No data-science degree required</div>
      <div class="guide-title">New to ML? Here's <span>everything in 60 seconds.</span></div>
      <div class="guide-desc">
        This page reads your sales data and runs <b>4 smart prediction engines</b> behind the scenes.
        Click any button below to dive into that model's full story — charts, tables, and a
        plain-English explanation of what it found.
      </div>
      <div class="guide-grid">
        <div class="guide-step">
          <div class="guide-step-num" style="background:linear-gradient(135deg,#EC4899,#8B5CF6);">1</div>
          <span class="guide-step-icon">🎯</span>
          <div class="guide-step-title">Payment Risk</div>
          <div class="guide-step-desc">Tells you <b>which customers are unlikely to pay</b> — so you call them first, before the money is lost.</div>
        </div>
        <div class="guide-step">
          <div class="guide-step-num" style="background:linear-gradient(135deg,#8B5CF6,#3B82F6);">2</div>
          <span class="guide-step-icon">📈</span>
          <div class="guide-step-title">Revenue Forecast</div>
          <div class="guide-step-desc">Predicts <b>how much money you'll likely make</b> in the next 7 days, based on your recent trend.</div>
        </div>
        <div class="guide-step">
          <div class="guide-step-num" style="background:linear-gradient(135deg,#3B82F6,#06B6D4);">3</div>
          <span class="guide-step-icon">📦</span>
          <div class="guide-step-title">Channel Intelligence</div>
          <div class="guide-step-desc">Shows <b>which sales channels are growing or shrinking</b> — so you know where to invest energy.</div>
        </div>
        <div class="guide-step">
          <div class="guide-step-num" style="background:linear-gradient(135deg,#06B6D4,#10B981);">4</div>
          <span class="guide-step-icon">👥</span>
          <div class="guide-step-title">Customer Segments</div>
          <div class="guide-step-desc">Groups customers into <b>VIP, Regular, and At Risk</b> — so you know who to protect and who to win back.</div>
        </div>
      </div>
      <div class="guide-footer-row">
        <div class="guide-hint">💬&nbsp; Don't want to read everything? Just ask the <b>AI Copilot</b> on the right in plain English instead.</div>
      </div>
    </div>
    """, unsafe_allow_html=True)
    gcol1, gcol2 = st.columns([1, 5])
    with gcol1:
        if st.button("Got it, hide this", key="dismiss_guide", use_container_width=True):
            st.session_state["ml_show_guide"] = False
            st.rerun()


# ══════════════════════════════════════════════════════════════════
# ② MODEL LAUNCHER — 4 big clickable cards (replaces "Active Models")
# ══════════════════════════════════════════════════════════════════
st.markdown("""
<div class="sec-block" style="margin-top:1rem;">
  <div class="sec-eyebrow">◈ &nbsp;CHOOSE WHAT TO EXPLORE<div class="line"></div></div>
  <div class="sec-title">What do you want to see?</div>
  <div class="sec-subtitle">Click a card to open that model's full breakdown — detailed charts, tables, and a plain-English explanation of what ARIA found.</div>
</div>
""", unsafe_allow_html=True)

active = st.session_state["ml_active_model"]

LAUNCHERS = [
    ("risk",     "🎯", "Payment Risk",        "lc-risk",     f"{n_at_risk} at-risk now",          "See who might not pay"),
    ("forecast", "📈", "Revenue Forecast",    "lc-forecast", f"₹{forecast_7d:,.0f} next 7 days",  "See where money is heading"),
    ("channel",  "📦", "Channel Intelligence","lc-channel",  f"{n_declining_channels} declining", "See what's growing or shrinking"),
    ("segment",  "👥", "Customer Segments",   "lc-segment",  f"{n_vip} VIP · {n_risk_cust} at risk", "See who matters most"),
]

lcols = st.columns(4)
for i, (key, icon, title, css_cls, stat, desc) in enumerate(LAUNCHERS):
    with lcols[i]:
        glow = "active-glow" if active == key else ""
        st.markdown(f"""
        <div class="launcher-card {css_cls} {glow}">
          <span class="launcher-icon">{icon}</span>
          <div class="launcher-title">{title}</div>
          <div class="launcher-desc">{desc}</div>
          <span class="launcher-stat">{stat}</span>
          <div class="launcher-cta">{'● Currently Viewing' if active == key else '○ Click to explore →'}</div>
        </div>
        """, unsafe_allow_html=True)
        if st.button(f"Open {title}", key=f"btn_{key}", use_container_width=True):
            st.session_state["ml_active_model"] = key
            st.rerun()

if active is not None:
    if st.button("← Back to Business Outlook", key="back_to_outlook"):
        st.session_state["ml_active_model"] = None
        st.rerun()


# ══════════════════════════════════════════════════════════════════
# MAIN CONTENT AREA — split into detail column + persistent chat dock
# ══════════════════════════════════════════════════════════════════
main_col, chat_col = st.columns([2.3, 1])

with main_col:

    # ──────────────────────────────────────────────────────────
    # OVERVIEW MODE — Business Outlook banner (default, no model open)
    # ──────────────────────────────────────────────────────────
    if active is None:
        outlook_cache_key = f"outlook-{cache_key}"
        with st.spinner("🔮 ARIA is writing your business briefing…"):
            outlook_text = generate_business_outlook(outlook_cache_key, ml_context_text)

        if not outlook_text:
            risk_word = "concerning" if n_at_risk > 5 else "manageable"
            trend_word = "growing" if ml_results.get("model2", {}).get("trend_slope", 0) >= 0 else "slowing down"
            outlook_text = (
                f"Based on your data, your revenue trend is currently <b>{trend_word}</b>, with a 7-day "
                f"forecast of ₹{forecast_7d:,.0f}. There {'are' if n_at_risk != 1 else 'is'} "
                f"<b>{n_at_risk} transaction(s)</b> flagged as payment risks worth ₹{at_risk_amount:,.0f} — "
                f"a {risk_word} amount to keep an eye on. {n_declining_channels} sales channel(s) are "
                f"trending down this month, and {n_risk_cust} customer(s) are showing signs of drifting away. "
                f"<b>Recommended action:</b> follow up on the highest-risk payments first, then check in "
                f"personally with any at-risk VIP customers. "
                f"<i>(Add a GROQ API key in .env for a fully personalised AI-written briefing.)</i>"
            )

        st.markdown(f"""
        <div class="model-banner mb-outlook">
          <div class="mb-head">
            <span style="font-size:1.4rem;">🌤️</span>
            <span class="mb-eyebrow" style="color:#34D399;">AI Business Briefing · Written just now</span>
          </div>
          <div class="mb-title">What ARIA thinks will happen next</div>
          <div class="mb-narrative">{outlook_text}</div>
          <div class="mb-pills">
            <div class="mb-pill">📅 <span>₹{forecast_7d:,.0f} expected next 7 days</span></div>
            <div class="mb-pill">⚠️ <span>{n_at_risk} payments at risk</span></div>
            <div class="mb-pill">📦 <span>{n_declining_channels} channel(s) declining</span></div>
            <div class="mb-pill">🌟 <span>{n_vip} VIP customers</span></div>
            <div class="mb-pill">🔻 <span>{n_risk_cust} customers drifting away</span></div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<div class="gc"><div class="gc-title"><div class="gc-dot" style="background:var(--gold);"></div>📖 Quick Glossary — What Is What?</div>', unsafe_allow_html=True)
        glossary_items = [
            ("🎯", "Risk Score", "A percentage (0-100%) showing how likely a customer is to NOT pay. Higher = more worrying.", "#EF4444"),
            ("📈", "Forecast", "ARIA's best guess at next week's revenue, based on your recent daily pattern.", "#8B5CF6"),
            ("📦", "Trend %", "How much a channel's sales changed between the first and second half of the month.", "#3B82F6"),
            ("👥", "Segment", "A label (VIP / Regular / At Risk) grouping similar customers together automatically.", "#F59E0B"),
            ("✅", "Accuracy", "How often the model's predictions matched reality when tested — higher is more trustworthy.", "#10B981"),
        ]
        gcols = st.columns(len(glossary_items))
        for gi, (icon, term, defn, color) in enumerate(glossary_items):
            with gcols[gi]:
                st.markdown(f"""
                <div class="gloss-row" style="border-bottom:none;">
                  <span class="gloss-icon">{icon}</span>
                  <div>
                    <div class="gloss-term" style="color:{color};">{term}</div>
                    <div class="gloss-def">{defn}</div>
                  </div>
                </div>
                """, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    # ──────────────────────────────────────────────────────────
    # MODEL 1 DETAIL — PAYMENT RISK (red/pink banner)
    # ──────────────────────────────────────────────────────────
    elif active == "risk":
        if "model1" in ml_results:
            m1 = ml_results["model1"]
            m1_df = m1["df"].copy()
            total_txns = len(m1_df)
            paid_txns = int(m1_df["IsPaid"].sum())
            at_risk = int((m1_df["RiskScore"] > 0.5).sum())
            at_risk_amt = float(m1_df[m1_df["RiskScore"] > 0.5]["Balance Due"].sum())
            acc_pct = m1["accuracy"] * 100

            narr_key = f"narr-risk-{cache_key}"
            with st.spinner("🔮 ARIA is analysing payment risk…"):
                narrative = generate_model_narrative(narr_key, "Payment Risk Predictor", ml_context_text)
            if not narrative:
                narrative = (
                    f"Out of <b>{total_txns} transactions</b>, ARIA flagged <b>{at_risk}</b> as likely to go "
                    f"unpaid, worth <b>₹{at_risk_amt:,.0f}</b> in outstanding dues. The model is right about "
                    f"<b>{acc_pct:.0f}%</b> of the time when tested against real outcomes. "
                    f"<b>Recommended action:</b> call the highest-risk customers first — every day of delay "
                    f"makes recovery harder."
                )

            st.markdown(f"""
            <div class="model-banner mb-risk">
              <div class="mb-head">
                <span style="font-size:1.4rem;">🎯</span>
                <span class="mb-eyebrow" style="color:#FCA5A5;">Model 1 · Payment Risk Predictor</span>
              </div>
              <div class="mb-title">Who's unlikely to pay — and how much is at stake</div>
              <div class="mb-narrative">{narrative}</div>
              <div class="mb-pills">
                <div class="mb-pill">🧾 <span>{total_txns} total transactions</span></div>
                <div class="mb-pill">✅ <span>{paid_txns} fully paid</span></div>
                <div class="mb-pill">⚠️ <span>{at_risk} flagged at risk</span></div>
                <div class="mb-pill">💰 <span>₹{at_risk_amt:,.0f} at stake</span></div>
                <div class="mb-pill">🎯 <span>{acc_pct:.0f}% model accuracy</span></div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            # Step-by-step "how ARIA got here"
            st.markdown("""
            <div class="sec-block" style="margin-top:1.2rem;">
              <div class="sec-eyebrow" style="color:#FCA5A5;">◈ &nbsp;HOW ARIA FIGURED THIS OUT<div class="line"></div></div>
              <div class="sec-title">Step-by-step: from raw data to risk score</div>
            </div>
            """, unsafe_allow_html=True)
            steps = [
                ("Looked at every past transaction", "ARIA studied who paid on time, who paid late, and who never paid — across every customer, salesperson, and channel in your data."),
                ("Found the patterns behind non-payment", "It learned which combinations (e.g. certain payment methods, certain channels, certain customers) are linked to unpaid bills."),
                ("Scored every current transaction 0–100%", "Each transaction got a risk score: low score = safe, high score = needs attention."),
                ("Tested itself before trusting itself", f"ARIA checked its own predictions against real outcomes it hadn't seen yet — it got this right {acc_pct:.0f}% of the time."),
            ]
            for i, (title, desc) in enumerate(steps):
                color = ["#EF4444","#F97316","#EC4899","#8B5CF6"][i]
                st.markdown(f"""
                <div class="step-card">
                  <div class="step-num" style="background:{color};">{i+1}</div>
                  <div class="step-body"><div class="step-title">{title}</div><div class="step-desc">{desc}</div></div>
                </div>
                """, unsafe_allow_html=True)

            # KPIs
            st.markdown("<div style='height:.8rem;'></div>", unsafe_allow_html=True)
            k1, k2, k3, k4, k5 = st.columns(5)
            with k1: st.markdown(f'<div class="kpi blue"><div class="kpi-label">Total Transactions</div><div class="kpi-value blue">{total_txns}</div><div class="kpi-sub">In dataset</div></div>', unsafe_allow_html=True)
            with k2: st.markdown(f'<div class="kpi green"><div class="kpi-label">Fully Paid</div><div class="kpi-value green">{paid_txns}</div><div class="kpi-sub">{paid_txns/total_txns*100:.1f}% of total</div></div>', unsafe_allow_html=True)
            with k3: st.markdown(f'<div class="kpi red"><div class="kpi-label">At Risk</div><div class="kpi-value red">{at_risk}</div><div class="kpi-sub">Risk score &gt; 50%</div></div>', unsafe_allow_html=True)
            with k4: st.markdown(f'<div class="kpi gold"><div class="kpi-label">Amount at Risk</div><div class="kpi-value gold">₹{at_risk_amt:,.0f}</div><div class="kpi-sub">Outstanding dues</div></div>', unsafe_allow_html=True)
            with k5:
                cls = "green" if acc_pct >= 70 else "gold"
                st.markdown(f'<div class="kpi {cls}"><div class="kpi-label">Model Accuracy</div><div class="kpi-value {cls}">{acc_pct:.1f}%</div><div class="kpi-sub">Test set score</div></div>', unsafe_allow_html=True)

            st.markdown("<div style='height:.8rem;'></div>", unsafe_allow_html=True)

            cf1, cf2, cf3 = st.columns(3)
            with cf1:
                risk_filter = st.selectbox("Show transactions", ["All", "At Risk Only (>50%)", "High Risk (>70%)", "Paid Only"], key="rf_filter")
            with cf2:
                sort_by = st.selectbox("Sort by", ["Risk Score (High→Low)", "Balance Due (High→Low)", "Total Amount"], key="rf_sort")
            with cf3:
                ch_filter = st.multiselect("Channel filter", options=sorted(m1_df["Sales Channel"].dropna().unique().tolist()), placeholder="All channels", key="rf_ch")

            disp = m1_df.copy()
            if risk_filter == "At Risk Only (>50%)": disp = disp[disp["RiskScore"] > 0.5]
            elif risk_filter == "High Risk (>70%)": disp = disp[disp["RiskScore"] > 0.7]
            elif risk_filter == "Paid Only": disp = disp[disp["IsPaid"] == 1]
            if ch_filter: disp = disp[disp["Sales Channel"].isin(ch_filter)]
            if sort_by == "Risk Score (High→Low)": disp = disp.sort_values("RiskScore", ascending=False)
            elif sort_by == "Balance Due (High→Low)": disp = disp.sort_values("Balance Due", ascending=False)
            else: disp = disp.sort_values("Total Amount", ascending=False)

            c_left, c_right = st.columns([1.6, 1])
            with c_left:
                st.markdown(f'<div class="gc"><div class="gc-title"><div class="gc-dot" style="background:var(--red);"></div>Transaction Risk List · {len(disp)} shown</div>', unsafe_allow_html=True)
                for _, row in disp.head(20).iterrows():
                    risk = float(row["RiskScore"]); is_paid = int(row["IsPaid"]); bal = float(row.get("Balance Due", 0))
                    party = str(row.get("Party Name", "—")); salesman = str(row.get("Salesman Name", "—"))
                    channel = str(row.get("Sales Channel", "—")); total = float(row.get("Total Amount", 0)); fu = str(row.get("Follow up date", "—"))
                    if is_paid == 1: badge = '<span class="risk-badge rb-paid">✓ PAID</span>'
                    elif risk > 0.7: badge = '<span class="risk-badge rb-risk">⚠ HIGH RISK</span>'
                    elif risk > 0.5: badge = '<span class="risk-badge rb-risk">⚠ AT RISK</span>'
                    else: badge = '<span class="risk-badge rb-partial">~ PARTIAL</span>'
                    risk_bar_w = int(risk * 100)
                    risk_color = "#EF4444" if risk > 0.7 else ("#F59E0B" if risk > 0.5 else "#10B981")
                    st.markdown(f"""
                    <div class="risk-row">
                      {badge}
                      <div style="flex:1;min-width:0;">
                        <div class="risk-party">{party}</div>
                        <div style="font-size:.65rem;color:var(--txt3);">{salesman} · {channel}</div>
                        <div style="height:3px;background:rgba(255,255,255,.05);border-radius:99px;margin-top:4px;overflow:hidden;">
                          <div style="height:100%;width:{risk_bar_w}%;background:{risk_color};border-radius:99px;"></div>
                        </div>
                      </div>
                      <div style="text-align:right;flex-shrink:0;">
                        <div class="risk-amt">₹{total:,.0f}</div>
                        <div class="risk-due">Due: ₹{bal:,.0f}</div>
                        <div style="font-size:.6rem;color:var(--txt3);">FU: {fu}</div>
                      </div>
                    </div>
                    """, unsafe_allow_html=True)
                if len(disp) > 20:
                    st.markdown(f"<div style='font-size:.72rem;color:var(--txt3);margin-top:.5rem;text-align:center;'>+{len(disp)-20} more transactions</div>", unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)

            with c_right:
                st.markdown('<div class="gc"><div class="gc-title"><div class="gc-dot" style="background:var(--red);"></div>Risk Distribution</div>', unsafe_allow_html=True)
                paid_c = int(m1_df["IsPaid"].sum()); risk_c = int((m1_df["RiskScore"] > 0.5).sum())
                safe_c = max(total_txns - paid_c - risk_c, 0)
                donut = go.Figure(go.Pie(values=[paid_c, safe_c, risk_c], labels=["Paid", "Low Risk", "At Risk"], hole=0.62,
                    marker=dict(colors=["#10B981","#3B82F6","#EF4444"], line=dict(color="rgba(0,0,0,0)", width=0)), textfont=dict(size=9)))
                style_fig(donut, height=200, showlegend=True, legend=dict(orientation="h", y=-0.1, font=dict(size=9)))
                st.plotly_chart(donut, use_container_width=True)
                st.markdown("</div>", unsafe_allow_html=True)

                st.markdown("<div style='height:.5rem;'></div>", unsafe_allow_html=True)
                st.markdown('<div class="gc"><div class="gc-title"><div class="gc-dot" style="background:var(--blue);"></div>What Drives Risk?</div>', unsafe_allow_html=True)
                fi = m1["feature_importance"]
                fi_sorted = sorted(fi.items(), key=lambda x: -x[1])[:6]
                FEAT_LABELS = {"Total Amount":"Invoice Amount","PaymentType_enc":"Payment Method","SalesChannel_enc":"Sales Channel",
                               "Salesman_enc":"Salesman","PartyName_enc":"Customer","DayOfWeek":"Day of Week","DayOfMonth":"Day of Month","FollowUpDays":"Follow-up Gap"}
                fi_max = fi_sorted[0][1] if fi_sorted else 1
                for idx, (fname, fval) in enumerate(fi_sorted):
                    pct = fval / fi_max * 100
                    lbl = FEAT_LABELS.get(fname, fname)
                    color = PAL[idx % len(PAL)]
                    st.markdown(f"""
                    <div style="display:flex;align-items:center;gap:8px;margin-bottom:5px;">
                      <div style="font-size:.72rem;color:var(--txt2);flex:1;">{lbl}</div>
                      <div style="flex:1;height:4px;border-radius:99px;background:rgba(255,255,255,.05);overflow:hidden;">
                        <div style="height:100%;width:{pct:.0f}%;background:{color};border-radius:99px;"></div>
                      </div>
                      <div style="font-family:'JetBrains Mono',monospace;font-size:.62rem;color:var(--txt3);flex:0 0 35px;text-align:right;">{fval:.2f}</div>
                    </div>
                    """, unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)

            st.markdown(f'<div class="insight-box red">🔮 <b>{at_risk}</b> transactions have a risk score above 50%, with <b>₹{at_risk_amt:,.0f}</b> at stake. Prioritise follow-up calls on the High Risk ones first — the model is <b>{acc_pct:.0f}% accurate</b> on held-out data.</div>', unsafe_allow_html=True)
        elif tx_df.empty:
            st.info("No Transactions sheet found in the uploaded file. This model requires a Transactions sheet with columns: Total Amount, Received/Paid Amount, Balance Due, Payment Type, Sales Channel, Salesman Name.")
        else:
            st.error(f"Model 1 error: {ml_results.get('model1_error', 'Unknown')}")

    # ──────────────────────────────────────────────────────────
    # MODEL 2 DETAIL — REVENUE FORECAST (purple/blue banner)
    # ──────────────────────────────────────────────────────────
    elif active == "forecast":
        if "model2" in ml_results:
            m2 = ml_results["model2"]
            historical, forecast, conf_lo, conf_hi = m2["historical"], m2["forecast"], m2["conf_lo"], m2["conf_hi"]
            slope, rmse = m2["trend_slope"], m2["rmse"]
            total_hist, total_fore = sum(historical), sum(forecast)
            avg_hist = total_hist / len(historical) if historical else 0
            avg_fore = sum(forecast) / len(forecast) if forecast else 0
            trend_dir = "📈 Growing" if slope > 0 else "📉 Declining"
            trend_color = "green" if slope > 0 else "red"

            narr_key = f"narr-forecast-{cache_key}"
            with st.spinner("🔮 ARIA is projecting your revenue…"):
                narrative = generate_model_narrative(narr_key, "Revenue Forecast", ml_context_text)
            if not narrative:
                narrative = (
                    f"Your revenue is currently <b>{'growing' if slope>0 else 'declining'}</b> at about "
                    f"₹{abs(slope):,.0f}/day. Based on this, ARIA expects <b>₹{total_fore:,.0f}</b> total over "
                    f"the next 7 days, compared to ₹{total_hist:,.0f} over the last {len(historical)} days. "
                    f"<b>Recommended action:</b> {'keep doing what is working — repeat your best days.' if slope>0 else 'investigate which channels or salespeople are slipping before the trend deepens.'}"
                )

            st.markdown(f"""
            <div class="model-banner mb-forecast">
              <div class="mb-head">
                <span style="font-size:1.4rem;">📈</span>
                <span class="mb-eyebrow" style="color:#C4B5FD;">Model 2 · Revenue Forecast Engine</span>
              </div>
              <div class="mb-title">Where your money is heading next</div>
              <div class="mb-narrative">{narrative}</div>
              <div class="mb-pills">
                <div class="mb-pill">📅 <span>₹{total_fore:,.0f} forecast (7 days)</span></div>
                <div class="mb-pill">📊 <span>₹{avg_fore:,.0f}/day average</span></div>
                <div class="mb-pill">{'📈' if slope>0 else '📉'} <span>₹{slope:+,.0f}/day trend</span></div>
                <div class="mb-pill">📏 <span>±₹{rmse:,.0f} typical error</span></div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown("""
            <div class="sec-block" style="margin-top:1.2rem;">
              <div class="sec-eyebrow" style="color:#C4B5FD;">◈ &nbsp;HOW ARIA FIGURED THIS OUT<div class="line"></div></div>
              <div class="sec-title">Step-by-step: from history to forecast</div>
            </div>
            """, unsafe_allow_html=True)
            steps2 = [
                ("Collected your daily sales history", f"ARIA pulled together {len(historical)} days of actual revenue from your daily report."),
                ("Drew the best-fit trend line", "It calculated the straight line that most closely follows your day-to-day ups and downs."),
                ("Projected that line 7 days forward", f"Following the trend, ARIA predicts ₹{total_fore:,.0f} total over the next week."),
                ("Added a confidence range", "Because the future isn't certain, each day's forecast comes with a realistic ±15% band rather than one rigid number."),
            ]
            for i, (title, desc) in enumerate(steps2):
                color = ["#8B5CF6","#3B82F6","#06B6D4","#A78BFA"][i]
                st.markdown(f"""
                <div class="step-card">
                  <div class="step-num" style="background:{color};">{i+1}</div>
                  <div class="step-body"><div class="step-title">{title}</div><div class="step-desc">{desc}</div></div>
                </div>
                """, unsafe_allow_html=True)

            st.markdown("<div style='height:.8rem;'></div>", unsafe_allow_html=True)
            kf1, kf2, kf3, kf4 = st.columns(4)
            with kf1: st.markdown(f'<div class="kpi purple"><div class="kpi-label">7-Day Forecast</div><div class="kpi-value purple">₹{total_fore:,.0f}</div><div class="kpi-sub">Projected total</div></div>', unsafe_allow_html=True)
            with kf2: st.markdown(f'<div class="kpi blue"><div class="kpi-label">Avg Forecast/Day</div><div class="kpi-value blue">₹{avg_fore:,.0f}</div><div class="kpi-sub">vs ₹{avg_hist:,.0f} historical</div></div>', unsafe_allow_html=True)
            with kf3:
                cls = "green" if slope > 0 else "red"
                st.markdown(f'<div class="kpi {cls}"><div class="kpi-label">Trend Direction</div><div class="kpi-value {cls}" style="font-size:1.1rem;">{trend_dir}</div><div class="kpi-sub">Slope: ₹{slope:+,.0f}/day</div></div>', unsafe_allow_html=True)
            with kf4: st.markdown(f'<div class="kpi gold"><div class="kpi-label">Forecast Error</div><div class="kpi-value gold">₹{rmse:,.0f}</div><div class="kpi-sub">RMSE on training data</div></div>', unsafe_allow_html=True)

            st.markdown("<div style='height:.8rem;'></div>", unsafe_allow_html=True)
            n_hist = len(historical)
            x_hist = list(range(1, n_hist + 1))
            x_fore = list(range(n_hist + 1, n_hist + len(forecast) + 1))

            fig_f = go.Figure()
            fig_f.add_trace(go.Scatter(x=x_hist, y=historical, name="Historical", mode="lines+markers",
                line=dict(color="#3B82F6", width=2.5, shape="spline"), marker=dict(size=4, color="#3B82F6")))
            fig_f.add_trace(go.Scatter(x=x_fore + x_fore[::-1], y=conf_hi + conf_lo[::-1], fill="toself",
                fillcolor="rgba(139,92,246,0.12)", line=dict(color="rgba(139,92,246,0)"), name="Confidence Band", showlegend=True))
            fig_f.add_trace(go.Scatter(x=x_fore, y=forecast, name="Forecast", mode="lines+markers",
                line=dict(color="#8B5CF6", width=2.5, dash="dot", shape="spline"), marker=dict(size=6, color="#8B5CF6", symbol="diamond")))
            fig_f.add_vline(x=n_hist + 0.5, line=dict(color="rgba(255,255,255,0.12)", width=1, dash="dash"),
                annotation_text="Forecast →", annotation_font=dict(color="#A78BFA", size=9))
            fig_f.update_xaxes(showgrid=False, tickfont=dict(size=9))
            fig_f.update_yaxes(showgrid=True, gridcolor="rgba(255,255,255,.04)", tickprefix="₹", tickformat=",.0f", tickfont=dict(size=9))
            style_fig(fig_f, height=360, showlegend=True, legend=dict(orientation="h", y=1.1, font=dict(size=10)))
            st.plotly_chart(fig_f, use_container_width=True)

            fc1, fc2 = st.columns(2)
            with fc1:
                st.markdown('<div class="gc"><div class="gc-title"><div class="gc-dot" style="background:var(--purple);"></div>7-Day Forecast Breakdown</div>', unsafe_allow_html=True)
                for i, (fval, lo, hi) in enumerate(zip(forecast, conf_lo, conf_hi)):
                    day_n = n_hist + i + 1
                    width_pct = min(int(fval / max(forecast) * 100), 100) if max(forecast) else 0
                    st.markdown(f"""
                    <div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;">
                      <div style="font-family:'JetBrains Mono',monospace;font-size:.6rem;color:var(--txt3);flex:0 0 50px;">Day {day_n}</div>
                      <div style="flex:1;height:6px;border-radius:99px;background:rgba(255,255,255,.05);overflow:hidden;">
                        <div style="height:100%;width:{width_pct}%;background:linear-gradient(90deg,#8B5CF6,#3B82F6);border-radius:99px;"></div>
                      </div>
                      <div style="font-family:'JetBrains Mono',monospace;font-size:.7rem;color:var(--txt1);flex:0 0 90px;text-align:right;">₹{fval:,.0f}</div>
                      <div style="font-size:.62rem;color:var(--txt3);flex:0 0 110px;text-align:right;">±₹{(hi-lo)/2:,.0f}</div>
                    </div>
                    """, unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)
            with fc2:
                st.markdown('<div class="gc"><div class="gc-title"><div class="gc-dot" style="background:var(--green);"></div>Historical Summary</div>', unsafe_allow_html=True)
                st.markdown(f"""
                <div style="display:flex;flex-direction:column;gap:.4rem;">
                  <div style="display:flex;justify-content:space-between;padding:.35rem 0;border-bottom:1px solid rgba(255,255,255,.04);">
                    <span style="font-size:.78rem;color:var(--txt3);">Days of history</span>
                    <span style="font-family:'JetBrains Mono',monospace;font-size:.75rem;color:var(--txt1);">{len(historical)}</span>
                  </div>
                  <div style="display:flex;justify-content:space-between;padding:.35rem 0;border-bottom:1px solid rgba(255,255,255,.04);">
                    <span style="font-size:.78rem;color:var(--txt3);">Total historical revenue</span>
                    <span style="font-family:'JetBrains Mono',monospace;font-size:.75rem;color:var(--txt1);">₹{total_hist:,.0f}</span>
                  </div>
                  <div style="display:flex;justify-content:space-between;padding:.35rem 0;border-bottom:1px solid rgba(255,255,255,.04);">
                    <span style="font-size:.78rem;color:var(--txt3);">Best historical day</span>
                    <span style="font-family:'JetBrains Mono',monospace;font-size:.75rem;color:#10B981;">₹{max(historical):,.0f}</span>
                  </div>
                  <div style="display:flex;justify-content:space-between;padding:.35rem 0;border-bottom:1px solid rgba(255,255,255,.04);">
                    <span style="font-size:.78rem;color:var(--txt3);">Worst historical day</span>
                    <span style="font-family:'JetBrains Mono',monospace;font-size:.75rem;color:#EF4444;">₹{min(historical):,.0f}</span>
                  </div>
                  <div style="display:flex;justify-content:space-between;padding:.35rem 0;">
                    <span style="font-size:.78rem;color:var(--txt3);">Projected next 7 days</span>
                    <span style="font-family:'JetBrains Mono',monospace;font-size:.75rem;color:#A78BFA;">₹{total_fore:,.0f}</span>
                  </div>
                </div>
                """, unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)

            trend_msg = f"Revenue is **{'growing' if slope>0 else 'declining'}** at ₹{abs(slope):,.0f}/day. Next 7 days projected at ₹{total_fore:,.0f} total."
            st.markdown(f'<div class="insight-box {trend_color}">📈 {trend_msg}</div>', unsafe_allow_html=True)
        else:
            err = ml_results.get("model2_error", "")
            st.info(f"Revenue forecast requires at least 5 days of historical data from the Daily Report sheet. {err}")

    # ──────────────────────────────────────────────────────────
    # MODEL 3 DETAIL — CHANNEL INTELLIGENCE (cyan/blue banner)
    # ──────────────────────────────────────────────────────────
    elif active == "channel":
        if "model3" in ml_results:
            ch_df = ml_results["model3"]["df"].copy()
            total_ch_rev = ch_df["Revenue"].sum()
            best_ch = ch_df.iloc[0]["Sales Channel"] if not ch_df.empty else "—"
            best_rev = float(ch_df.iloc[0]["Revenue"]) if not ch_df.empty else 0
            n_growing = len(ch_df[ch_df["TrendPct"] > 0])
            best_coll_idx = ch_df["CollectionRate"].idxmax() if not ch_df.empty else 0
            best_coll_ch = ch_df.loc[best_coll_idx, "Sales Channel"] if not ch_df.empty else "—"
            best_coll_rt = float(ch_df.loc[best_coll_idx, "CollectionRate"]) * 100 if not ch_df.empty else 0

            narr_key = f"narr-channel-{cache_key}"
            with st.spinner("🔮 ARIA is comparing your channels…"):
                narrative = generate_model_narrative(narr_key, "Channel Intelligence", ml_context_text)
            if not narrative:
                narrative = (
                    f"<b>{best_ch}</b> is your top channel with ₹{best_rev:,.0f} in revenue. "
                    f"{n_growing} of {len(ch_df)} channels are growing this month. "
                    f"<b>{best_coll_ch}</b> collects payment best, at {best_coll_rt:.0f}%. "
                    f"<b>Recommended action:</b> double down on what's working in {best_ch}, and investigate "
                    f"any channel showing a declining trend before it slips further."
                )

            st.markdown(f"""
            <div class="model-banner mb-channel">
              <div class="mb-head">
                <span style="font-size:1.4rem;">📦</span>
                <span class="mb-eyebrow" style="color:#67E8F9;">Model 3 · Channel Intelligence</span>
              </div>
              <div class="mb-title">Where your revenue really comes from</div>
              <div class="mb-narrative">{narrative}</div>
              <div class="mb-pills">
                <div class="mb-pill">🏆 <span>{best_ch} leads</span></div>
                <div class="mb-pill">📈 <span>{n_growing} of {len(ch_df)} growing</span></div>
                <div class="mb-pill">💳 <span>{best_coll_ch} best collector</span></div>
                <div class="mb-pill">📊 <span>₹{total_ch_rev:,.0f} total revenue</span></div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown("""
            <div class="sec-block" style="margin-top:1.2rem;">
              <div class="sec-eyebrow" style="color:#67E8F9;">◈ &nbsp;HOW ARIA FIGURED THIS OUT<div class="line"></div></div>
              <div class="sec-title">Step-by-step: comparing your channels</div>
            </div>
            """, unsafe_allow_html=True)
            steps3 = [
                ("Grouped every transaction by channel", "ARIA sorted all your sales into the channel they came from — direct, online, exhibitions, and so on."),
                ("Measured revenue and order patterns", "For each channel, it calculated total revenue, transaction count, and average order size."),
                ("Split the month in half to spot trends", "By comparing the first 15 days vs the last 15 days, ARIA can tell which channels are accelerating and which are slowing down."),
                ("Checked who actually pays", "It also measured the collection rate per channel — high revenue doesn't matter if the money never arrives."),
            ]
            for i, (title, desc) in enumerate(steps3):
                color = ["#06B6D4","#3B82F6","#10B981","#0EA5E9"][i]
                st.markdown(f"""
                <div class="step-card">
                  <div class="step-num" style="background:{color};">{i+1}</div>
                  <div class="step-body"><div class="step-title">{title}</div><div class="step-desc">{desc}</div></div>
                </div>
                """, unsafe_allow_html=True)

            st.markdown("<div style='height:.8rem;'></div>", unsafe_allow_html=True)
            ck1, ck2, ck3, ck4 = st.columns(4)
            with ck1: st.markdown(f'<div class="kpi purple"><div class="kpi-label">Top Channel</div><div class="kpi-value purple" style="font-size:1rem;">{best_ch[:12]}</div><div class="kpi-sub">₹{best_rev:,.0f} revenue</div></div>', unsafe_allow_html=True)
            with ck2: st.markdown(f'<div class="kpi green"><div class="kpi-label">Growing Channels</div><div class="kpi-value green">{n_growing}</div><div class="kpi-sub">of {len(ch_df)} total</div></div>', unsafe_allow_html=True)
            with ck3: st.markdown(f'<div class="kpi blue"><div class="kpi-label">Best Collection Rate</div><div class="kpi-value blue" style="font-size:1rem;">{best_coll_ch[:12]}</div><div class="kpi-sub">{best_coll_rt:.1f}% paid on time</div></div>', unsafe_allow_html=True)
            with ck4: st.markdown(f'<div class="kpi gold"><div class="kpi-label">Total Channels</div><div class="kpi-value gold">{len(ch_df)}</div><div class="kpi-sub">Active channels</div></div>', unsafe_allow_html=True)

            st.markdown("<div style='height:.8rem;'></div>", unsafe_allow_html=True)
            ch_left, ch_right = st.columns([1.4, 1])
            CH_ICONS = {"Telesales":"📞","Office":"🏢","Medical Store":"💊","Direct":"🤝","Amazon":"📦","Flipkart":"🛒","Other":"📡"}

            with ch_left:
                st.markdown('<div class="gc"><div class="gc-title"><div class="gc-dot" style="background:var(--cyan);"></div>Channel Breakdown + Trend</div>', unsafe_allow_html=True)
                for ridx, row in ch_df.iterrows():
                    ch_name = row["Sales Channel"]; rev = float(row["Revenue"]); txns = int(row["Transactions"])
                    avg_ord = float(row["AvgOrder"]); coll_rate = float(row["CollectionRate"]) * 100; trend_pct = float(row["TrendPct"])
                    pct_share = rev / total_ch_rev * 100 if total_ch_rev else 0
                    icon = CH_ICONS.get(ch_name, "📡")
                    arrow = "▲" if trend_pct > 0 else "▼"
                    t_color = "#10B981" if trend_pct > 0 else "#EF4444"
                    bar_color = PAL[list(ch_df["Sales Channel"]).index(ch_name) % len(PAL)]
                    st.markdown(f"""
                    <div class="ch-trend-row">
                      <span class="ch-icon">{icon}</span>
                      <div style="flex:1;min-width:0;">
                        <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:3px;">
                          <span class="ch-name">{ch_name}</span>
                          <span style="font-family:'JetBrains Mono',monospace;font-size:.7rem;color:{t_color};">{arrow} {abs(trend_pct):.1f}%</span>
                        </div>
                        <div style="height:5px;background:rgba(255,255,255,.05);border-radius:99px;overflow:hidden;">
                          <div style="height:100%;width:{pct_share:.1f}%;background:{bar_color};border-radius:99px;"></div>
                        </div>
                        <div style="display:flex;justify-content:space-between;margin-top:3px;">
                          <span style="font-size:.6rem;color:var(--txt3);">{txns} txns · ₹{avg_ord:,.0f} avg</span>
                          <span style="font-size:.6rem;color:var(--txt3);">{coll_rate:.0f}% collected · ₹{rev:,.0f}</span>
                        </div>
                      </div>
                    </div>
                    """, unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)

            with ch_right:
                st.markdown('<div class="gc"><div class="gc-title"><div class="gc-dot" style="background:var(--blue);"></div>Revenue Share</div>', unsafe_allow_html=True)
                donut_ch = go.Figure(go.Pie(values=ch_df["Revenue"].tolist(), labels=ch_df["Sales Channel"].tolist(), hole=0.58,
                    marker=dict(colors=PAL[:len(ch_df)], line=dict(color="rgba(0,0,0,0)", width=0)), textfont=dict(size=9)))
                style_fig(donut_ch, height=220, showlegend=True, legend=dict(orientation="v", font=dict(size=8.5), x=1.0))
                st.plotly_chart(donut_ch, use_container_width=True)
                st.markdown("</div>", unsafe_allow_html=True)

                st.markdown("<div style='height:.5rem;'></div>", unsafe_allow_html=True)
                st.markdown('<div class="gc"><div class="gc-title"><div class="gc-dot" style="background:var(--green);"></div>Collection Rate by Channel</div>', unsafe_allow_html=True)
                fig_coll = go.Figure(go.Bar(x=ch_df["CollectionRate"] * 100, y=ch_df["Sales Channel"], orientation="h",
                    marker=dict(color=[("#10B981" if v > 0.6 else ("#F59E0B" if v > 0.4 else "#EF4444")) for v in ch_df["CollectionRate"]])))
                fig_coll.update_xaxes(showgrid=True, gridcolor="rgba(255,255,255,.04)", ticksuffix="%", range=[0, 110])
                fig_coll.update_yaxes(showgrid=False)
                style_fig(fig_coll, height=200, showlegend=False)
                st.plotly_chart(fig_coll, use_container_width=True)
                st.markdown("</div>", unsafe_allow_html=True)

            declining = ch_df[ch_df["TrendPct"] < 0]
            if not declining.empty:
                worst = declining.iloc[declining["TrendPct"].argmin()]
                st.markdown(f'<div class="insight-box red">⚠️ <b>{worst["Sales Channel"]}</b> is the most declining channel ({worst["TrendPct"]:.1f}% drop second half vs first half). Investigate and consider reactivation strategy.</div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="insight-box green">✅ All channels are trending upward. <b>{best_ch}</b> leads with ₹{best_rev:,.0f} in revenue ({best_rev/total_ch_rev*100:.1f}% share).</div>', unsafe_allow_html=True)
        else:
            st.info(f"Channel intelligence requires transaction data. Error: {ml_results.get('model3_error','No data')}")

    # ──────────────────────────────────────────────────────────
    # MODEL 4 DETAIL — CUSTOMER SEGMENTS (gold/amber banner)
    # ──────────────────────────────────────────────────────────
    elif active == "segment":
        if "model4" in ml_results:
            cust_df = ml_results["model4"]["df"].copy()
            total_cust = len(cust_df)
            vip_count = int((cust_df["Segment"] == "VIP").sum())
            reg_count = int((cust_df["Segment"] == "Regular").sum())
            risk_count = int((cust_df["Segment"] == "At Risk").sum())
            vip_rev = float(cust_df[cust_df["Segment"] == "VIP"]["Total_Revenue"].sum())

            narr_key = f"narr-segment-{cache_key}"
            with st.spinner("🔮 ARIA is grouping your customers…"):
                narrative = generate_model_narrative(narr_key, "Customer Segmentation", ml_context_text)
            if not narrative:
                narrative = (
                    f"ARIA grouped your <b>{total_cust} customers</b> into 3 segments: <b>{vip_count} VIP</b> "
                    f"customers worth ₹{vip_rev:,.0f} combined, <b>{reg_count} Regular</b> customers with room "
                    f"to grow, and <b>{risk_count} At Risk</b> customers who may be drifting away. "
                    f"<b>Recommended action:</b> protect your VIPs with priority service, and reach out "
                    f"personally to anyone in the At Risk group before you lose them."
                )

            st.markdown(f"""
            <div class="model-banner mb-segment">
              <div class="mb-head">
                <span style="font-size:1.4rem;">👥</span>
                <span class="mb-eyebrow" style="color:#FCD34D;">Model 4 · Customer Segmentation</span>
              </div>
              <div class="mb-title">Who matters most — and who needs attention</div>
              <div class="mb-narrative">{narrative}</div>
              <div class="mb-pills">
                <div class="mb-pill">🌟 <span>{vip_count} VIP customers</span></div>
                <div class="mb-pill">📊 <span>{reg_count} Regular customers</span></div>
                <div class="mb-pill">⚠️ <span>{risk_count} At Risk</span></div>
                <div class="mb-pill">💰 <span>₹{vip_rev:,.0f} VIP revenue</span></div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown("""
            <div class="sec-block" style="margin-top:1.2rem;">
              <div class="sec-eyebrow" style="color:#FCD34D;">◈ &nbsp;HOW ARIA FIGURED THIS OUT<div class="line"></div></div>
              <div class="sec-title">Step-by-step: grouping similar customers</div>
            </div>
            """, unsafe_allow_html=True)
            steps4 = [
                ("Measured every customer on 6 dimensions", "Order count, total revenue, average order size, payment reliability, outstanding balance, and how many channels they buy through."),
                ("Put similar customers in the same group", "Using a clustering technique, ARIA automatically found 3 natural groups — without being told what to look for."),
                ("Ranked the groups by value", "The group with the highest average revenue became 'VIP', the lowest became 'At Risk', and the middle became 'Regular'."),
                ("Attached a recommended action to each group", "Every segment comes with a specific next step — so you know exactly how to treat each type of customer."),
            ]
            for i, (title, desc) in enumerate(steps4):
                color = ["#F59E0B","#EC4899","#8B5CF6","#FBBF24"][i]
                st.markdown(f"""
                <div class="step-card">
                  <div class="step-num" style="background:{color};">{i+1}</div>
                  <div class="step-body"><div class="step-title">{title}</div><div class="step-desc">{desc}</div></div>
                </div>
                """, unsafe_allow_html=True)

            st.markdown("<div style='height:.8rem;'></div>", unsafe_allow_html=True)
            sk1, sk2, sk3, sk4 = st.columns(4)
            with sk1: st.markdown(f'<div class="kpi purple"><div class="kpi-label">Total Customers</div><div class="kpi-value purple">{total_cust}</div><div class="kpi-sub">Unique parties</div></div>', unsafe_allow_html=True)
            with sk2: st.markdown(f'<div class="kpi gold"><div class="kpi-label">VIP Customers</div><div class="kpi-value gold">{vip_count}</div><div class="kpi-sub">₹{vip_rev:,.0f} combined</div></div>', unsafe_allow_html=True)
            with sk3: st.markdown(f'<div class="kpi blue"><div class="kpi-label">Regular</div><div class="kpi-value blue">{reg_count}</div><div class="kpi-sub">Growth opportunity</div></div>', unsafe_allow_html=True)
            with sk4:
                cls = "red" if risk_count > 0 else "green"
                st.markdown(f'<div class="kpi {cls}"><div class="kpi-label">At Risk</div><div class="kpi-value {cls}">{risk_count}</div><div class="kpi-sub">Need attention</div></div>', unsafe_allow_html=True)

            st.markdown("<div style='height:.8rem;'></div>", unsafe_allow_html=True)
            seg_left, seg_right = st.columns([1.3, 1])
            SEG_ACTIONS = {
                "VIP": ("🌟", "#F59E0B", "sl-vip", "Priority relationship. Offer loyalty discounts, early access to new products, dedicated account manager. Protect at all costs."),
                "Regular": ("📊", "#60A5FA", "sl-regular", "Growth potential. Upsell complementary products, increase order frequency with targeted promotions and follow-up calls."),
                "At Risk": ("⚠️", "#EF4444", "sl-risk", "Immediate action needed. Personal outreach, understand pain points, offer flexible payment terms to retain the relationship."),
            }
            with seg_left:
                for seg in ["VIP", "Regular", "At Risk"]:
                    seg_rows = cust_df[cust_df["Segment"] == seg]
                    if seg_rows.empty: continue
                    icon, color, badge_cls, action = SEG_ACTIONS[seg]
                    avg_rev = float(seg_rows["Total_Revenue"].mean()); avg_txns = float(seg_rows["Total_Transactions"].mean())
                    avg_pay = float(seg_rows["Payment_Rate"].mean()) * 100
                    st.markdown(f"""
                    <div class="seg-card">
                      <div class="seg-card-top">
                        <div style="display:flex;align-items:center;gap:8px;">
                          <span style="font-size:1.2rem;">{icon}</span><span class="seg-name">{seg} Customers</span>
                        </div>
                        <span class="seg-label {badge_cls}">{len(seg_rows)} customer{'s' if len(seg_rows)!=1 else ''}</span>
                      </div>
                      <div style="font-size:.72rem;color:var(--txt3);margin-bottom:.6rem;">{', '.join(seg_rows['Party Name'].tolist())}</div>
                      <div class="seg-stats">
                        <div class="seg-stat"><div class="seg-stat-val" style="color:{color};">₹{avg_rev:,.0f}</div><div class="seg-stat-lbl">Avg Revenue</div></div>
                        <div class="seg-stat"><div class="seg-stat-val" style="color:{color};">{avg_txns:.0f}</div><div class="seg-stat-lbl">Avg Orders</div></div>
                        <div class="seg-stat"><div class="seg-stat-val" style="color:{color};">{avg_pay:.0f}%</div><div class="seg-stat-lbl">Pay Rate</div></div>
                      </div>
                      <div class="seg-action">💡 <b>Action:</b> {action}</div>
                    </div>
                    """, unsafe_allow_html=True)

            with seg_right:
                st.markdown('<div class="gc"><div class="gc-title"><div class="gc-dot" style="background:var(--gold);"></div>Revenue by Segment</div>', unsafe_allow_html=True)
                seg_rev = cust_df.groupby("Segment")["Total_Revenue"].sum().reset_index()
                seg_colors = {"VIP":"#F59E0B","Regular":"#3B82F6","At Risk":"#EF4444"}
                fig_seg = go.Figure(go.Pie(values=seg_rev["Total_Revenue"].tolist(), labels=seg_rev["Segment"].tolist(), hole=0.55,
                    marker=dict(colors=[seg_colors.get(s, "#8B5CF6") for s in seg_rev["Segment"]], line=dict(color="rgba(0,0,0,0)", width=0)), textfont=dict(size=9.5)))
                style_fig(fig_seg, height=220, showlegend=True, legend=dict(orientation="h", y=-0.1, font=dict(size=9)))
                st.plotly_chart(fig_seg, use_container_width=True)
                st.markdown("</div>", unsafe_allow_html=True)

                st.markdown("<div style='height:.5rem;'></div>", unsafe_allow_html=True)
                st.markdown('<div class="gc"><div class="gc-title"><div class="gc-dot" style="background:var(--green);"></div>Payment Rate by Segment</div>', unsafe_allow_html=True)
                seg_pay = cust_df.groupby("Segment")["Payment_Rate"].mean().reset_index()
                fig_pay = go.Figure(go.Bar(x=seg_pay["Segment"], y=seg_pay["Payment_Rate"] * 100,
                    marker=dict(color=[seg_colors.get(s, "#8B5CF6") for s in seg_pay["Segment"]], line=dict(color="rgba(0,0,0,0)", width=0))))
                fig_pay.update_yaxes(showgrid=True, gridcolor="rgba(255,255,255,.04)", ticksuffix="%", range=[0,110])
                fig_pay.update_xaxes(showgrid=False)
                style_fig(fig_pay, height=175, showlegend=False)
                st.plotly_chart(fig_pay, use_container_width=True)
                st.markdown("</div>", unsafe_allow_html=True)

            if risk_count > 0:
                risk_names = cust_df[cust_df["Segment"] == "At Risk"]["Party Name"].tolist()
                st.markdown(f'<div class="insight-box red">⚠️ <b>{", ".join(risk_names)}</b> {"is" if len(risk_names)==1 else "are"} in the At Risk segment. Schedule personal outreach this week — recovering an existing customer costs 5× less than acquiring a new one.</div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="insight-box green">✅ All customers are in VIP or Regular segments. Focus on moving Regular customers to VIP by increasing their order frequency and offering bulk incentives.</div>', unsafe_allow_html=True)
        else:
            st.info(f"Customer segmentation requires transaction data. Error: {ml_results.get('model4_error','No data')}")


# ══════════════════════════════════════════════════════════════════
# PERSISTENT AI CHAT DOCK — right column, always visible
# ══════════════════════════════════════════════════════════════════
with chat_col:
    scope_labels = {
        None: "Business Outlook (overview)",
        "risk": "Payment Risk model",
        "forecast": "Revenue Forecast model",
        "channel": "Channel Intelligence model",
        "segment": "Customer Segmentation model",
    }
    scope_label = scope_labels.get(active)

    st.markdown('<div class="chat-wrap">', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="chat-head">
      <span style="font-size:1.2rem;">🤖</span>
      <span class="chat-title">ARIA Copilot</span>
      <span class="chat-badge">LIVE</span>
    </div>
    <div class="chat-scope-tag">📍 Scoped to: {scope_label}</div>
    <div class="chat-desc">Ask about future predictions, risk, forecasts, channels, or customers — ARIA answers using the real model results.</div>
    """, unsafe_allow_html=True)

    if st.session_state["ml_chat_history"]:
        for turn in st.session_state["ml_chat_history"][-8:]:
            st.markdown(f"""
            <div class="chat-bubble user"><span class="chat-bubble-label">You</span>{turn['q']}</div>
            <div class="chat-bubble ai"><span class="chat-bubble-label">🔮 ARIA</span>{turn['a']}</div>
            """, unsafe_allow_html=True)
    else:
        st.markdown('<div class="chat-empty">No questions yet — try a suggestion below, or type your own future-focused question.</div>', unsafe_allow_html=True)

    # Context-aware suggestions based on which model is open
    suggestion_map = {
        None: ["Will I hit my sales target this week?", "What's the single biggest risk right now?"],
        "risk": ["Which customers should I call first today?", "How much money is realistically recoverable?"],
        "forecast": ["Why is my revenue trend going this direction?", "What would it take to grow 20% next week?"],
        "channel": ["Which channel should I invest more in?", "Why is my weakest channel struggling?"],
        "segment": ["How do I turn Regular customers into VIPs?", "What should I say to At Risk customers?"],
    }
    suggestions = suggestion_map.get(active, suggestion_map[None])

    clicked_suggestion = None
    for si, sugg in enumerate(suggestions):
        if st.button(sugg, key=f"chat_sugg_{active}_{si}", use_container_width=True):
            clicked_suggestion = sugg

    user_q = st.chat_input("Ask ARIA anything…", key=f"chat_input_{active}")
    final_question = clicked_suggestion or user_q

    if final_question:
        with st.spinner("🔮 ARIA is thinking…"):
            answer = generate_chat_reply(final_question, ml_context_text, st.session_state["ml_chat_history"], scope_label)
        st.session_state["ml_chat_history"].append({"q": final_question, "a": answer})
        st.rerun()

    if st.session_state["ml_chat_history"]:
        if st.button("🗑️ Clear conversation", key="clear_chat", use_container_width=True):
            st.session_state["ml_chat_history"] = []
            st.rerun()

    st.markdown("</div>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# FOOTER + NAVIGATION
# ══════════════════════════════════════════════════════════════════
st.markdown("<div style='height:1.5rem;'></div>", unsafe_allow_html=True)
fn1, fn2, fn3 = st.columns(3)
with fn1:
    if st.button("← Dashboard", use_container_width=True, key="foot_dash"):
        st.switch_page("pages/1_Dashboard.py")
with fn2:
    if st.button("🧠 AI Copilot (Full Page)", use_container_width=True, key="foot_chat"):
        st.switch_page("pages/2_AI_Chat.py")
with fn3:
    if st.button("📤 New Upload", use_container_width=True, key="foot_upload"):
        st.switch_page("app.py")

st.markdown("""
<div class="aria-footer">
  ARIA · ML Intelligence Engine · v3.0 · scikit-learn + Random Forest + K-Means · Neural Glass
</div>
""", unsafe_allow_html=True)