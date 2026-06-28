import streamlit as st
if not st.session_state.get("password_correct", False):
    st.stop()

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import os
import requests
from datetime import datetime, date

from utils.data_processor import (
    get_salesperson_totals,
    get_zero_sales,
    get_daily_totals,
    get_category_totals,
    get_mtd_total,
    get_ytd_total,
    get_bank_balance,
    get_cash_balance,
    get_bill_count,
    get_returns_total,
    get_product_category_totals,
    get_data_summary_for_ai,
    get_date_columns,
)
from utils.data_loader import get_raw_dataframe, get_clean_dataframe

st.set_page_config(
    page_title="ARIA · Dashboard",
    page_icon="◆",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────
# LOAD DATA
# ─────────────────────────────────────────────────────────────────
if "df" not in st.session_state:
    st.warning("No dataset loaded.")
    st.page_link("app.py", label="← Upload Data", icon="⬅️")
    st.stop()

df       = st.session_state["df"]
filename = st.session_state.get("filename", "dataset.xlsx")

raw_df = pd.DataFrame()
if "uploaded_file_bytes" in st.session_state:
    try:
        raw_df = get_raw_dataframe(BytesIO(st.session_state["uploaded_file_bytes"]))
    except Exception:
        pass

tx_df = pd.DataFrame()
if "uploaded_file_bytes" in st.session_state:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(
            BytesIO(st.session_state["uploaded_file_bytes"]), data_only=True
        )
        if "Transactions" in wb.sheetnames:
            ws   = wb["Transactions"]
            rows = list(ws.iter_rows(values_only=True))
            hdrs = rows[0]
            tx_df = pd.DataFrame(
                [dict(zip(hdrs, r)) for r in rows[1:] if any(v is not None for v in r)]
            )
    except Exception:
        pass

# ─────────────────────────────────────────────────────────────────
# CHART THEME
# ─────────────────────────────────────────────────────────────────
PLOT_BASE = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="Inter, sans-serif", color="#94A3B8", size=11),
)
BLUE_PAL = ["#3B82F6","#06B6D4","#8B5CF6","#10B981","#F59E0B","#EF4444","#F97316","#84CC16"]

def style_fig(fig, height=300, margin=None, **kwargs):
    fig.update_layout(**PLOT_BASE, height=height, **kwargs)
    fig.update_layout(margin=margin or dict(l=8, r=8, t=32, b=8))
    return fig

# ─────────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;600&family=Space+Grotesk:wght@400;500;600;700&display=swap');
*,*::before,*::after{box-sizing:border-box;}
:root{
  --bg:#020510;--surface:rgba(255,255,255,0.032);
  --border:rgba(255,255,255,0.072);--border2:rgba(255,255,255,0.12);
  --blue:#3B82F6;--blue2:#60A5FA;--cyan:#06B6D4;--purple:#8B5CF6;
  --green:#10B981;--gold:#F59E0B;--red:#EF4444;--orange:#F97316;
  --txt1:#F1F5F9;--txt2:#94A3B8;--txt3:#475569;
}
html,body,.stApp{
  background:
    radial-gradient(ellipse 90% 55% at 15% -10%,rgba(59,130,246,0.10) 0%,transparent 60%),
    radial-gradient(ellipse 70% 50% at 85% 105%,rgba(139,92,246,0.08) 0%,transparent 60%),
    #020510 !important;
  color:var(--txt2);font-family:'Inter',sans-serif;
}
#MainMenu,footer,header{visibility:hidden;}
.block-container{padding:5.5rem 2.4vw 3rem !important;max-width:100% !important;}

/* TOPBAR */
#topbar{
  position:fixed;top:0;left:0;right:0;z-index:300;
  display:flex;align-items:center;justify-content:space-between;
  padding:.75rem 2rem;
  background:rgba(2,5,16,0.88);
  border-bottom:1px solid var(--border);
  backdrop-filter:blur(32px);
}
.tb-brand{display:flex;align-items:center;gap:.7rem;}
.tb-mark{
  width:32px;height:32px;border-radius:9px;
  background:linear-gradient(135deg,#3B82F6,#8B5CF6 55%,#06B6D4);
  display:flex;align-items:center;justify-content:center;
  font-size:.78rem;font-weight:900;color:#fff;
  box-shadow:0 0 20px rgba(59,130,246,.45);
}
.tb-name{font-family:'Space Grotesk',sans-serif;font-size:1rem;font-weight:700;color:var(--txt1);}
.tb-sep{color:var(--border2);margin:0 .3rem;}
.tb-page{font-size:.82rem;color:var(--txt2);}
.tb-right{display:flex;align-items:center;gap:.55rem;}
.tb-tag{
  font-family:'JetBrains Mono',monospace;font-size:.56rem;
  padding:3px 11px;border-radius:6px;
  background:var(--surface);border:1px solid var(--border);color:var(--txt3);
}
.tb-live{
  display:flex;align-items:center;gap:5px;
  font-family:'JetBrains Mono',monospace;font-size:.56rem;
  padding:3px 11px;border-radius:6px;
  background:rgba(16,185,129,.07);border:1px solid rgba(16,185,129,.22);color:var(--green);
}
.live-dot{width:5px;height:5px;border-radius:50%;background:var(--green);box-shadow:0 0 7px var(--green);animation:pulse 2s infinite;}
@keyframes pulse{0%,100%{opacity:1;transform:scale(1);}50%{opacity:.3;transform:scale(.8);}}

/* SIDEBAR */
section[data-testid="stSidebar"]{
  background:rgba(3,8,22,0.97) !important;
  border-right:1px solid var(--border) !important;
  backdrop-filter:blur(24px) !important;
}
section[data-testid="stSidebar"] *{color:var(--txt2) !important;}
.sb-brand{display:flex;align-items:center;gap:.6rem;padding:.8rem .9rem .7rem;border-bottom:1px solid var(--border);margin-bottom:.6rem;}
.sb-mark{width:34px;height:34px;border-radius:9px;background:linear-gradient(135deg,#3B82F6,#8B5CF6 55%,#06B6D4);
  display:flex;align-items:center;justify-content:center;font-size:.82rem;font-weight:900;color:#fff;
  box-shadow:0 0 16px rgba(59,130,246,.35);flex-shrink:0;}
.sb-name{font-family:'Space Grotesk',sans-serif;font-size:.92rem;font-weight:700;color:var(--txt1) !important;}
.sb-ver{font-family:'JetBrains Mono',monospace;font-size:.48rem;color:var(--txt3) !important;}
.sb-sec{display:flex;align-items:center;gap:8px;padding:.45rem .9rem .2rem;margin-top:.2rem;}
.sb-sec-label{font-family:'JetBrains Mono',monospace;font-size:.5rem;letter-spacing:2.5px;text-transform:uppercase;color:var(--blue2) !important;font-weight:600;}
.sb-sec-line{flex:1;height:1px;background:linear-gradient(90deg,rgba(59,130,246,.25),transparent);}
.sb-status{margin:.2rem .9rem .5rem;display:flex;align-items:center;gap:7px;
  background:rgba(16,185,129,.06);border:1px solid rgba(16,185,129,.2);border-radius:10px;padding:.38rem .75rem;}
.sb-status-dot{width:6px;height:6px;border-radius:50%;background:var(--green);box-shadow:0 0 7px var(--green);animation:pulse 2s infinite;}
.sb-status-txt{font-family:'JetBrains Mono',monospace;font-size:.55rem;color:var(--green) !important;letter-spacing:.5px;}
.sb-stats{padding:.1rem .9rem .3rem;}
.sb-stat{display:flex;justify-content:space-between;align-items:center;padding:.2rem 0;border-bottom:1px solid rgba(255,255,255,.03);}
.sb-stat:last-child{border-bottom:none;}
.sb-stat-l{font-size:.72rem;color:var(--txt3) !important;}
.sb-stat-r{font-family:'JetBrains Mono',monospace;font-size:.7rem;color:var(--txt1) !important;font-weight:600;}
.sb-date-chip{margin:.15rem .9rem .1rem;background:rgba(59,130,246,.07);border:1px solid rgba(59,130,246,.18);
  border-radius:10px;padding:.5rem .75rem;font-family:'JetBrains Mono',monospace;font-size:.6rem;
  color:var(--blue2) !important;display:flex;align-items:center;justify-content:space-between;}
.sb-date-chip b{color:var(--txt1) !important;}
.nav-section-title{font-family:'JetBrains Mono',monospace;font-size:.48rem;letter-spacing:2.2px;text-transform:uppercase;
  color:var(--blue2) !important;padding:.5rem .9rem .2rem;display:flex;align-items:center;gap:7px;margin-top:.3rem;}
.nav-section-title::after{content:'';flex:1;height:1px;background:linear-gradient(90deg,rgba(59,130,246,.2),transparent);}
.nav-btn{display:flex;align-items:center;gap:9px;margin:.12rem .7rem;padding:.42rem .75rem;border-radius:10px;
  background:rgba(255,255,255,.025);border:1px solid rgba(255,255,255,.055);cursor:pointer;transition:all .2s ease;text-decoration:none;}
.nav-btn:hover{background:rgba(59,130,246,.09);border-color:rgba(59,130,246,.22);transform:translateX(3px);}
.nav-btn-icon{font-size:.85rem;flex-shrink:0;}
.nav-btn-txt{font-size:.75rem;color:var(--txt2) !important;font-weight:500;}
.nav-btn-arrow{margin-left:auto;font-size:.6rem;color:var(--txt3) !important;}
.sb-prog-wrap{padding:.1rem .9rem .4rem;}
.sb-prog-track{height:6px;border-radius:99px;background:rgba(255,255,255,.06);overflow:hidden;margin:.35rem 0 .2rem;}
.sb-prog-fill{height:100%;border-radius:99px;background:linear-gradient(90deg,var(--blue),var(--purple),var(--cyan));}
.sb-prog-cap{display:flex;justify-content:space-between;font-size:.65rem;color:var(--txt3) !important;}

/* SECTION HEADERS */
.sec-block{margin:2.8rem 0 1.2rem;}
.sec-eyebrow{display:flex;align-items:center;gap:10px;font-family:'JetBrains Mono',monospace;font-size:.58rem;
  color:var(--blue2);letter-spacing:2.5px;text-transform:uppercase;margin-bottom:.5rem;}
.sec-eyebrow .line{flex:1;height:1px;background:linear-gradient(90deg,rgba(59,130,246,.3),transparent);}
.sec-title{font-family:'Space Grotesk',sans-serif;font-size:1.45rem;font-weight:700;color:var(--txt1);line-height:1.2;margin-bottom:.4rem;letter-spacing:-.5px;}
.sec-subtitle{font-size:.86rem;color:var(--txt3);line-height:1.6;max-width:720px;padding-bottom:.5rem;border-bottom:1px solid rgba(255,255,255,.04);margin-bottom:.2rem;}

/* HERO */
.hero{border-radius:22px;border:1px solid rgba(59,130,246,.2);
  background:linear-gradient(135deg,rgba(59,130,246,.09),rgba(139,92,246,.06),rgba(6,182,212,.06));
  backdrop-filter:blur(20px);padding:2rem 2.2rem 1.8rem;position:relative;overflow:hidden;margin-bottom:2rem;}
.hero::before{content:'';position:absolute;top:-80px;right:-60px;width:400px;height:400px;border-radius:50%;
  background:radial-gradient(circle,rgba(59,130,246,.12),transparent 70%);pointer-events:none;}
.hero-inner{position:relative;z-index:1;}
.hero-status{display:flex;align-items:center;gap:8px;font-family:'JetBrains Mono',monospace;font-size:.58rem;
  color:var(--green);letter-spacing:2px;text-transform:uppercase;margin-bottom:.8rem;}
.hero-sdot{width:6px;height:6px;border-radius:50%;background:var(--green);box-shadow:0 0 8px var(--green);animation:pulse 2s infinite;}
.hero-title{font-family:'Space Grotesk',sans-serif;font-size:2rem;font-weight:700;color:var(--txt1);margin-bottom:.5rem;letter-spacing:-.6px;line-height:1.15;}
.hero-title span{background:linear-gradient(135deg,#60A5FA,#8B5CF6,#06B6D4);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;}
.hero-desc{font-size:.88rem;color:var(--txt3);line-height:1.65;max-width:560px;margin-bottom:1.4rem;}
.hero-chips{display:flex;flex-wrap:wrap;gap:.5rem;}
.hero-chip{display:flex;align-items:center;gap:6px;font-family:'JetBrains Mono',monospace;font-size:.58rem;
  padding:5px 13px;border-radius:9px;background:rgba(255,255,255,.03);border:1px solid var(--border);color:var(--txt3);}
.hero-chip b{color:var(--txt1);}
.hero-health{position:absolute;top:1.8rem;right:2rem;z-index:1;font-family:'JetBrains Mono',monospace;
  font-size:.65rem;padding:8px 18px;border-radius:99px;font-weight:700;letter-spacing:.4px;}
.hg{background:rgba(16,185,129,.1);color:var(--green);border:1px solid rgba(16,185,129,.3);}
.hw{background:rgba(245,158,11,.1);color:var(--gold);border:1px solid rgba(245,158,11,.3);}
.hb{background:rgba(239,68,68,.1);color:var(--red);border:1px solid rgba(239,68,68,.3);}

/* KPI CARDS */
.kpi{border-radius:18px;border:1px solid var(--border);background:var(--surface);backdrop-filter:blur(16px);
  padding:1.2rem 1.3rem 1.1rem;position:relative;overflow:hidden;transition:transform .25s ease,border-color .25s ease,box-shadow .25s ease;}
.kpi::before{content:'';position:absolute;top:0;left:0;right:0;height:2.5px;border-radius:18px 18px 0 0;}
.kpi.blue::before{background:linear-gradient(90deg,#3B82F6,#06B6D4);}
.kpi.cyan::before{background:linear-gradient(90deg,#06B6D4,#3B82F6);}
.kpi.green::before{background:linear-gradient(90deg,#10B981,#06B6D4);}
.kpi.gold::before{background:linear-gradient(90deg,#F59E0B,#F97316);}
.kpi.purple::before{background:linear-gradient(90deg,#8B5CF6,#3B82F6);}
.kpi.red::before{background:linear-gradient(90deg,#EF4444,#F97316);}
.kpi:hover{transform:translateY(-4px);border-color:var(--border2);box-shadow:0 12px 40px rgba(59,130,246,.12);}
.kpi-icon-row{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:.7rem;}
.kpi-icon{font-size:1.3rem;}
.kpi-hint{font-size:.58rem;color:var(--txt3);cursor:help;border:1px solid var(--border);border-radius:50%;
  width:17px;height:17px;display:flex;align-items:center;justify-content:center;}
.kpi-label{font-family:'JetBrains Mono',monospace;font-size:.52rem;color:var(--txt3);text-transform:uppercase;letter-spacing:1.4px;margin-bottom:.35rem;}
.kpi-value{font-family:'Space Grotesk',sans-serif;font-size:1.7rem;font-weight:700;line-height:1;margin-bottom:.3rem;letter-spacing:-.5px;}
.kpi-value.blue{color:#60A5FA;}.kpi-value.cyan{color:#06B6D4;}.kpi-value.green{color:#10B981;}
.kpi-value.gold{color:#F59E0B;}.kpi-value.purple{color:#8B5CF6;}.kpi-value.red{color:#EF4444;}
.kpi-sub{font-size:.7rem;color:var(--txt3);line-height:1.4;}
.kpi-delta{display:inline-flex;align-items:center;gap:3px;font-size:.68rem;font-weight:700;margin-top:.3rem;}
.up{color:var(--green);}.dn{color:var(--red);}

/* AI CARD */
.ai-card{border-radius:20px;border:1px solid rgba(139,92,246,.25);
  background:linear-gradient(135deg,rgba(139,92,246,.08),rgba(59,130,246,.04),rgba(6,182,212,.04));
  padding:1.6rem 1.8rem;position:relative;overflow:hidden;}
.ai-card::before{content:'';position:absolute;top:-60px;right:-40px;width:260px;height:260px;border-radius:50%;
  background:radial-gradient(circle,rgba(139,92,246,.14),transparent 70%);pointer-events:none;}
.ai-card-header{display:flex;align-items:center;gap:10px;margin-bottom:1.1rem;position:relative;z-index:1;}
.ai-card-title{font-family:'Space Grotesk',sans-serif;font-size:1.1rem;font-weight:700;color:var(--txt1);}
.ai-badge{font-family:'JetBrains Mono',monospace;font-size:.52rem;padding:3px 10px;border-radius:99px;
  background:rgba(139,92,246,.12);border:1px solid rgba(139,92,246,.3);color:var(--purple);margin-left:auto;}
.ai-body{font-size:.9rem;color:var(--txt2);line-height:2;white-space:pre-wrap;position:relative;z-index:1;}

/* GLASS CARD */
.gc{border-radius:18px;border:1px solid var(--border);background:var(--surface);backdrop-filter:blur(16px);padding:1.3rem 1.4rem;height:100%;}
.gc-title{font-family:'JetBrains Mono',monospace;font-size:.56rem;color:var(--txt3);text-transform:uppercase;letter-spacing:1.6px;margin-bottom:1rem;display:flex;align-items:center;gap:7px;}
.gc-dot{width:6px;height:6px;border-radius:50%;flex-shrink:0;}
.gc-note{font-size:.73rem;color:var(--txt3);margin-top:.75rem;line-height:1.6;border-top:1px solid rgba(255,255,255,.04);padding-top:.65rem;}

/* LEADERBOARD */
.lb-row{display:flex;align-items:center;gap:10px;padding:.55rem 0;border-bottom:1px solid rgba(255,255,255,.035);}
.lb-row:last-child{border-bottom:none;}
.lb-rank{width:24px;font-family:'JetBrains Mono',monospace;font-size:.68rem;color:var(--txt3);text-align:center;flex-shrink:0;}
.lb-rank.r1{color:var(--gold);font-weight:700;}.lb-rank.r2{color:#C0C0C0;font-weight:600;}.lb-rank.r3{color:var(--orange);font-weight:600;}
.lb-av{width:32px;height:32px;border-radius:50%;flex-shrink:0;display:flex;align-items:center;justify-content:center;
  font-size:.65rem;font-weight:700;background:rgba(59,130,246,.12);color:var(--blue2);border:1px solid rgba(59,130,246,.2);}
.lb-av.top{background:rgba(245,158,11,.1);color:var(--gold);border-color:rgba(245,158,11,.22);}
.lb-info{flex:1;min-width:0;}
.lb-row-top{display:flex;justify-content:space-between;align-items:center;}
.lb-name{font-size:.82rem;color:var(--txt2);font-weight:500;}
.lb-cat{font-family:'JetBrains Mono',monospace;font-size:.5rem;color:var(--txt3);}
.lb-amount{font-family:'JetBrains Mono',monospace;font-size:.78rem;color:var(--txt1);font-weight:700;}
.lb-bar-track{height:3px;border-radius:99px;background:rgba(255,255,255,.05);margin-top:5px;overflow:hidden;}
.lb-bar-fill{height:100%;border-radius:99px;background:linear-gradient(90deg,var(--blue),var(--cyan));}
.tag-watch{font-family:'JetBrains Mono',monospace;font-size:.46rem;padding:1px 7px;border-radius:5px;margin-left:.4rem;
  background:rgba(245,158,11,.08);color:var(--gold);border:1px solid rgba(245,158,11,.2);}
.tag-ok{font-family:'JetBrains Mono',monospace;font-size:.46rem;padding:1px 7px;border-radius:5px;margin-left:.4rem;
  background:rgba(16,185,129,.07);color:var(--green);border:1px solid rgba(16,185,129,.18);}

/* CHANNEL ROWS */
.ch-row{display:flex;align-items:center;gap:.7rem;padding:.45rem 0;border-bottom:1px solid rgba(255,255,255,.03);}
.ch-row:last-child{border-bottom:none;}
.ch-name-block{min-width:0;flex:0 0 34%;}
.ch-name{font-size:.8rem;color:var(--txt2);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.ch-desc{font-size:.6rem;color:var(--txt3);display:block;margin-top:1px;}
.ch-bar-wrap{flex:1;height:5px;border-radius:99px;background:rgba(255,255,255,.05);overflow:hidden;}
.ch-bar{height:100%;border-radius:99px;}
.ch-val-block{text-align:right;flex:0 0 auto;}
.ch-val{font-family:'JetBrains Mono',monospace;font-size:.72rem;color:var(--txt1);font-weight:600;white-space:nowrap;}
.ch-pct{font-size:.6rem;color:var(--txt3);}

/* ALERT ROW */
.alert-row{display:flex;align-items:center;gap:9px;padding:.42rem .75rem;border-radius:11px;
  background:rgba(239,68,68,.05);border:1px solid rgba(239,68,68,.14);margin-bottom:.32rem;font-size:.77rem;color:var(--txt2);}
.alert-dot{width:6px;height:6px;border-radius:50%;background:var(--red);box-shadow:0 0 6px var(--red);flex-shrink:0;}

/* PROGRESS */
.prog-track{height:9px;border-radius:99px;background:rgba(255,255,255,.06);overflow:hidden;margin:.6rem 0 .35rem;}
.prog-fill{height:100%;border-radius:99px;background:linear-gradient(90deg,var(--blue),var(--purple),var(--cyan));}
.prog-cap{display:flex;justify-content:space-between;font-size:.7rem;color:var(--txt3);}

/* INSIGHT CALLOUT */
.insight-box{border-radius:13px;padding:.9rem 1.1rem;margin-top:.8rem;font-size:.8rem;line-height:1.7;color:var(--txt2);}
.insight-box.blue{background:rgba(59,130,246,.06);border:1px solid rgba(59,130,246,.15);}
.insight-box.gold{background:rgba(245,158,11,.06);border:1px solid rgba(245,158,11,.15);}
.insight-box.green{background:rgba(16,185,129,.06);border:1px solid rgba(16,185,129,.15);}
.insight-box.red{background:rgba(239,68,68,.06);border:1px solid rgba(239,68,68,.15);}

/* AI EXPLAIN BOX */
.ai-explain{border-radius:14px;padding:1rem 1.2rem;margin-top:1rem;background:rgba(139,92,246,.05);border:1px solid rgba(139,92,246,.15);}
.ai-explain-header{display:flex;align-items:center;gap:8px;margin-bottom:.7rem;font-family:'JetBrains Mono',monospace;font-size:.55rem;color:#A78BFA;text-transform:uppercase;letter-spacing:1.5px;}
.ai-bullet{display:flex;align-items:flex-start;gap:8px;margin-bottom:.45rem;}
.ai-bullet-icon{flex-shrink:0;margin-top:1px;}
.ai-bullet-txt{font-size:.8rem;color:#94A3B8;line-height:1.6;}
.ai-bullet-txt b{color:#F1F5F9;}

/* FOLLOW-UP ROWS */
.fu-row{display:flex;align-items:center;gap:9px;padding:.42rem .7rem;border-radius:10px;
  background:rgba(245,158,11,.04);border:1px solid rgba(245,158,11,.12);margin-bottom:.3rem;}
.fu-dot{width:5px;height:5px;border-radius:50%;background:var(--gold);flex-shrink:0;}
.fu-party{font-size:.76rem;color:var(--txt2);}
.fu-due{font-size:.62rem;color:var(--red);}

/* COLL ROWS */
.coll-row{display:flex;justify-content:space-between;align-items:center;padding:.4rem 0;border-bottom:1px solid rgba(255,255,255,.035);}
.coll-row:last-child{border-bottom:none;}
.coll-lbl{font-size:.78rem;color:var(--txt3);}
.coll-val{font-family:'JetBrains Mono',monospace;font-size:.8rem;color:var(--txt1);font-weight:600;}

/* CAT BAR CARDS */
.cat-bar-card{border-radius:16px;border:1px solid rgba(255,255,255,.06);background:rgba(255,255,255,.02);
  padding:.9rem 1.1rem;margin-bottom:.5rem;transition:all .2s;}
.cat-bar-card:hover{background:rgba(255,255,255,.04);border-color:rgba(255,255,255,.1);transform:translateX(4px);}
.cat-bar-card-top{display:flex;align-items:center;justify-content:space-between;margin-bottom:.55rem;}
.cat-bar-name{font-size:.85rem;font-weight:600;color:#F1F5F9;display:flex;align-items:center;gap:8px;}
.cat-bar-right{text-align:right;}
.cat-bar-val{font-family:'JetBrains Mono',monospace;font-size:.85rem;font-weight:700;}
.cat-bar-pct{font-size:.62rem;color:#475569;margin-top:1px;}
.cat-bar-track{height:7px;border-radius:99px;background:rgba(255,255,255,.05);overflow:hidden;}
.cat-bar-fill{height:100%;border-radius:99px;}

/* RETURN FLOW */
.flow-steps{display:flex;align-items:center;gap:.5rem;margin:1.2rem 0 .8rem;flex-wrap:wrap;}
.flow-step{flex:1;min-width:140px;border-radius:14px;padding:.9rem 1rem;text-align:center;}
.flow-step.gross{background:rgba(59,130,246,.08);border:1px solid rgba(59,130,246,.18);}
.flow-step.ret{background:rgba(239,68,68,.07);border:1px solid rgba(239,68,68,.16);}
.flow-step.net{background:rgba(16,185,129,.08);border:1px solid rgba(16,185,129,.18);}
.flow-step-icon{font-size:1.5rem;margin-bottom:.3rem;}
.flow-step-label{font-family:'JetBrains Mono',monospace;font-size:.5rem;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:.3rem;}
.flow-step-label.gross{color:#60A5FA;}.flow-step-label.ret{color:#EF4444;}.flow-step-label.net{color:#10B981;}
.flow-step-val{font-family:'Space Grotesk',sans-serif;font-size:1.2rem;font-weight:700;}
.flow-step-val.gross{color:#60A5FA;}.flow-step-val.ret{color:#EF4444;}.flow-step-val.net{color:#10B981;}
.flow-step-sub{font-size:.65rem;color:#475569;margin-top:2px;}
.flow-arrow{font-size:1.3rem;color:#475569;flex-shrink:0;}

/* EXPORT CARDS */
.export-card{border-radius:16px;border:1px solid var(--border);background:var(--surface);padding:1.2rem 1.2rem .9rem;transition:all .22s;}
.export-card:hover{border-color:var(--border2);box-shadow:0 6px 24px rgba(59,130,246,.08);}
.export-icon{font-size:1.5rem;margin-bottom:.5rem;}
.export-title{font-family:'Space Grotesk',sans-serif;font-size:.88rem;font-weight:600;color:var(--txt1);margin-bottom:.25rem;}
.export-desc{font-size:.72rem;color:var(--txt3);line-height:1.5;margin-bottom:.8rem;}

/* TABS */
div[data-testid="stTabs"] button{font-family:'JetBrains Mono',monospace !important;font-size:.6rem !important;letter-spacing:1px !important;text-transform:uppercase !important;color:var(--txt3) !important;}
div[data-testid="stTabs"] button[aria-selected="true"]{color:var(--blue2) !important;}

/* INPUTS */
div[data-testid="stSelectbox"] label,div[data-testid="stMultiSelect"] label,
div[data-testid="stDateInput"] label,div[data-testid="stNumberInput"] label{
  font-family:'JetBrains Mono',monospace !important;font-size:.53rem !important;
  text-transform:uppercase !important;letter-spacing:1.2px !important;color:var(--txt3) !important;}

/* BUTTONS */
.stButton button,.stDownloadButton button{
  background:linear-gradient(135deg,rgba(59,130,246,.12),rgba(139,92,246,.09)) !important;
  border:1px solid rgba(59,130,246,.25) !important;color:var(--txt1) !important;
  border-radius:11px !important;font-family:'Inter',sans-serif !important;
  font-weight:600 !important;transition:all .22s !important;font-size:.82rem !important;}
.stButton button:hover,.stDownloadButton button:hover{
  border-color:var(--cyan) !important;box-shadow:0 0 20px rgba(6,182,212,.12) !important;transform:translateY(-1px) !important;}

/* DATAFRAME */
div[data-testid="stDataFrame"]{border-radius:14px;overflow:hidden;border:1px solid var(--border);}

/* SCROLLBAR */
::-webkit-scrollbar{width:4px;height:4px;}
::-webkit-scrollbar-track{background:transparent;}
::-webkit-scrollbar-thumb{background:rgba(59,130,246,.3);border-radius:99px;}

/* FOOTER */
.aria-footer{text-align:center;margin-top:3rem;padding-top:1.5rem;border-top:1px solid var(--border);
  font-family:'JetBrains Mono',monospace;font-size:.58rem;color:var(--txt3);letter-spacing:1.2px;}

html{scroll-behavior:smooth;}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────
def section(title, subtitle, anchor_id=None):
    anchor = f'<span id="{anchor_id}"></span>' if anchor_id else ""
    st.markdown(f"""
    {anchor}
    <div class="sec-block">
      <div class="sec-eyebrow">◈ &nbsp;{title.upper()}<div class="line"></div></div>
      <div class="sec-title">{title}</div>
      <div class="sec-subtitle">{subtitle}</div>
    </div>
    """, unsafe_allow_html=True)

def insight(text, color="blue"):
    st.markdown(f'<div class="insight-box {color}">💡 {text}</div>', unsafe_allow_html=True)

def gc_open(dot_color="var(--blue)", title=""):
    st.markdown(
        f'<div class="gc"><div class="gc-title"><div class="gc-dot" style="background:{dot_color};"></div>{title}</div>',
        unsafe_allow_html=True,
    )

def gc_close(note=""):
    if note:
        st.markdown(f'<div class="gc-note">💡 {note}</div>', unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
# DATE COLUMNS
# ─────────────────────────────────────────────────────────────────
date_cols_all = get_date_columns(df)
n_cols        = len(date_cols_all)

def parse_col_date(col_name):
    try:
        return pd.to_datetime(str(col_name)).date()
    except Exception:
        return None

col_dates      = [parse_col_date(c) for c in date_cols_all]
has_real_dates = all(d is not None for d in col_dates)


# ─────────────────────────────────────────────────────────────────
# TOPBAR
# ─────────────────────────────────────────────────────────────────
st.markdown(f"""
<div id="topbar">
  <div class="tb-brand">
    <div class="tb-mark">A</div>
    <div class="tb-name">ARIA</div>
    <div class="tb-sep">·</div>
    <div class="tb-page">Sales Intelligence Dashboard</div>
  </div>
  <div class="tb-right">
    <div class="tb-tag">{filename}</div>
    <div class="tb-tag">LLaMA 3.3-70B</div>
    <div class="tb-live"><div class="live-dot"></div>Live</div>
  </div>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
# SIDEBAR — all filter state collected FIRST
# ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div class="sb-brand">
      <div class="sb-mark">A</div>
      <div>
        <div class="sb-name">ARIA Dashboard</div>
        <div class="sb-ver">v5.0 · NEURAL GLASS</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="sb-sec"><div class="sb-sec-label">System</div><div class="sb-sec-line"></div></div>', unsafe_allow_html=True)
    st.markdown('<div class="sb-status"><div class="sb-status-dot"></div><div class="sb-status-txt">ALL SYSTEMS ONLINE</div></div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="sb-stats">
      <div class="sb-stat"><span class="sb-stat-l">File</span><span class="sb-stat-r" style="font-size:.58rem;">{filename[:16]}…</span></div>
      <div class="sb-stat"><span class="sb-stat-l">Rows</span><span class="sb-stat-r">{len(df):,}</span></div>
      <div class="sb-stat"><span class="sb-stat-l">Sellers</span><span class="sb-stat-r">{len(df)}</span></div>
    </div>
    """, unsafe_allow_html=True)

    # ── DATE RANGE ──────────────────────────────────────────────
    st.markdown('<div class="sb-sec" style="margin-top:.5rem;"><div class="sb-sec-label">Date Range</div><div class="sb-sec-line"></div></div>', unsafe_allow_html=True)

    if has_real_dates:
        min_date = col_dates[0]
        max_date = col_dates[-1]
        d_from = st.date_input("From", value=min_date, min_value=min_date, max_value=max_date, key="date_from")
        d_to   = st.date_input("To",   value=max_date, min_value=min_date, max_value=max_date, key="date_to")
        d_from = min(d_from, d_to)
        lo = next((i + 1 for i, d in enumerate(col_dates) if d >= d_from), 1)
        hi = next((n_cols - i for i, d in enumerate(reversed(col_dates)) if d <= d_to), n_cols)
        hi = max(lo, hi)
        lo_label = col_dates[lo - 1].strftime("%d %b %Y")
        hi_label = col_dates[hi - 1].strftime("%d %b %Y")
        st.markdown(f"""
        <div class="sb-date-chip">
          <span>{lo_label} → {hi_label}</span><b>{hi - lo + 1}d</b>
        </div>""", unsafe_allow_html=True)
    else:
        preset = st.selectbox("Quick Select", ["Full Range","Last 7 Days","Last 14 Days","Last 30 Days"], key="preset")
        if preset == "Last 7 Days":    lo, hi = max(1, n_cols - 6), n_cols
        elif preset == "Last 14 Days": lo, hi = max(1, n_cols - 13), n_cols
        elif preset == "Last 30 Days": lo, hi = max(1, n_cols - 29), n_cols
        else:                          lo, hi = 1, n_cols
        lo, hi = st.slider("Range", 1, n_cols, (lo, hi), key="day_slider")
        st.markdown(f'<div class="sb-date-chip"><span>Day {lo} → {hi}</span><b>{hi - lo + 1}d</b></div>', unsafe_allow_html=True)

    # ── CHANNEL FILTER ──────────────────────────────────────────
    st.markdown('<div class="sb-sec" style="margin-top:.5rem;"><div class="sb-sec-label">Filters</div><div class="sb-sec-line"></div></div>', unsafe_allow_html=True)

    # Build full people list to get all categories
    all_people_df = get_salesperson_totals(df)
    all_cats = sorted(all_people_df["Category"].unique().tolist()) if not all_people_df.empty else []

    with st.expander("📡 Sales Channels", expanded=True):
        sel_cats = []
        for cat in all_cats:
            if st.checkbox(cat, value=True, key=f"cat_{cat}"):
                sel_cats.append(cat)

    # ── SALESPERSON FILTER ──────────────────────────────────────
    # Get names available under selected categories
    if sel_cats and not all_people_df.empty:
        avail_names = all_people_df[all_people_df["Category"].isin(sel_cats)]["Salesperson"].tolist()
    else:
        avail_names = []

    with st.expander("👤 Salespeople", expanded=False):
        if avail_names:
            col_a, col_b = st.columns(2)
            with col_a:
                if st.button("✓ All", use_container_width=True, key="sel_all"):
                    for n in avail_names:
                        st.session_state[f"sp_{n}"] = True
            with col_b:
                if st.button("✗ None", use_container_width=True, key="sel_none"):
                    for n in avail_names:
                        st.session_state[f"sp_{n}"] = False

            sel_people = []
            for name in avail_names:
                key = f"sp_{name}"
                if key not in st.session_state:
                    st.session_state[key] = True
                checked = st.checkbox(name, value=st.session_state[key], key=key)
                if checked:
                    sel_people.append(name)
        else:
            st.markdown("<div style='font-size:.72rem;color:var(--txt3);'>Select a channel first.</div>", unsafe_allow_html=True)
            sel_people = []

    # ── TARGET ──────────────────────────────────────────────────
    st.markdown('<div class="sb-sec" style="margin-top:.5rem;"><div class="sb-sec-label">Target</div><div class="sb-sec-line"></div></div>', unsafe_allow_html=True)
    monthly_target = st.number_input("Monthly Target (₹)", min_value=0, value=2_000_000, step=100_000)
    show_raw = st.checkbox("Show data tables", value=True)
    show_tx  = st.checkbox("Show transactions", value=True)

    # ── NAVIGATION ──────────────────────────────────────────────
    st.markdown('<div class="nav-section-title">Quick Navigation</div>', unsafe_allow_html=True)
    nav_items = [
        ("🧠", "AI Insights",         "#sec-ai"),
        ("📊", "Key Metrics",         "#sec-metrics"),
        ("📈", "Daily Trend",         "#sec-trend"),
        ("📡", "Sales Channels",      "#sec-channels"),
        ("🏆", "Team Performance",    "#sec-team"),
        ("🧬", "Product Categories",  "#sec-products"),
        ("💸", "Returns & Net Sales", "#sec-returns"),
        ("💳", "Transactions",        "#sec-transactions"),
        ("⬇️", "Export & Tables",     "#sec-export"),
    ]
    for icon, label, anchor in nav_items:
        st.markdown(f"""
        <a href="{anchor}" class="nav-btn" style="text-decoration:none;">
          <span class="nav-btn-icon">{icon}</span>
          <span class="nav-btn-txt">{label}</span>
          <span class="nav-btn-arrow">›</span>
        </a>
        """, unsafe_allow_html=True)

    st.markdown("<div style='height:.4rem;'></div>", unsafe_allow_html=True)
    if st.button("← New Upload", use_container_width=True):
        st.switch_page("app.py")
    if st.button("🔮 ML Intelligence →", use_container_width=True):
        st.switch_page("pages/3_ML_Intelligence.py")


# ─────────────────────────────────────────────────────────────────
# DATA COMPUTATION — ALL downstream data filtered consistently
# ─────────────────────────────────────────────────────────────────
selected_date_cols = date_cols_all[lo - 1: hi]

# Base filtered DataFrame (by date only — used for daily totals)
base_df = df[["Name"] + list(selected_date_cols)].copy()
base_df.columns = ["Name"] + list(selected_date_cols)

# ── Apply category + person filters to get filtered_df ──────────
# We filter the raw df rows by matching Names that belong to
# selected categories and selected people
all_pt = get_salesperson_totals(base_df)  # has Salesperson, Category, Total

# Determine which names pass the filters
if sel_cats:
    names_in_cats = all_pt[all_pt["Category"].isin(sel_cats)]["Salesperson"].tolist()
else:
    names_in_cats = []

if sel_people:
    names_final = [n for n in names_in_cats if n in sel_people]
else:
    names_final = names_in_cats

# filtered_df = only rows for selected people, with date cols
if names_final:
    filtered_df = base_df[base_df["Name"].isin(names_final)].copy()
else:
    # No selection → show empty but don't crash
    filtered_df = base_df.iloc[0:0].copy()

# ── Derived metrics all from filtered_df ────────────────────────
person_totals = get_salesperson_totals(filtered_df)
zero_list     = get_zero_sales(filtered_df)
daily_totals  = get_daily_totals(filtered_df)   # sums across filtered people per day
cat_totals    = get_category_totals(filtered_df)

# MTD / YTD / Bank / Cash from RAW df (global, not filtered)
real_mtd      = get_mtd_total(raw_df)    if not raw_df.empty else float(daily_totals.sum())
real_ytd      = get_ytd_total(raw_df)    if not raw_df.empty else 0.0
bank_balance  = get_bank_balance(raw_df) if not raw_df.empty else 0.0
cash_balance  = get_cash_balance(raw_df) if not raw_df.empty else 0.0
bill_count    = get_bill_count(raw_df)   if not raw_df.empty else "—"
returns_total = get_returns_total(raw_df, lo, hi) if not raw_df.empty else 0.0

total_sales   = float(person_totals["Total"].sum())     if not person_totals.empty else 0.0
avg_daily     = float(daily_totals.mean())               if len(daily_totals) else 0.0
n_active      = int((person_totals["Total"] > 0).sum())  if not person_totals.empty else 0
top_performer = person_totals.iloc[0]                    if not person_totals.empty else None

today_val    = float(daily_totals.iloc[-1])  if len(daily_totals) >= 1 else 0.0
yest_val     = float(daily_totals.iloc[-2])  if len(daily_totals) >= 2 else 0.0
day_chg_pct  = ((today_val - yest_val) / yest_val * 100) if yest_val else 0.0
week_this    = float(daily_totals.iloc[-7:].sum())    if len(daily_totals) >= 7  else float(daily_totals.sum())
week_prev    = float(daily_totals.iloc[-14:-7].sum()) if len(daily_totals) >= 14 else 0.0
week_chg_pct = ((week_this - week_prev) / week_prev * 100) if week_prev else 0.0
progress_pct = min(real_mtd / monthly_target * 100, 100) if monthly_target else 0

best_idx  = int(daily_totals.values.argmax()) if len(daily_totals) else 0
worst_idx = int(daily_totals.values.argmin()) if len(daily_totals) else 0

tp_name = top_performer["Salesperson"] if top_performer is not None else "—"
tp_val  = float(top_performer["Total"]) if top_performer is not None else 0.0

net_sales = max(total_sales - returns_total, 0)

w_cls, w_arrow = ("up", "▲") if week_chg_pct >= 0 else ("dn", "▼")

if week_chg_pct >= 5:    health_cls, health_txt = "hg", "🟢 Trending Up"
elif week_chg_pct <= -5: health_cls, health_txt = "hb", "🔴 Sales Declining"
else:                     health_cls, health_txt = "hw", "🟡 Holding Steady"

# Product categories — computed from filtered names within raw_df
prod_cats_raw = get_product_category_totals(raw_df, lo, hi) if not raw_df.empty else {}
# If people filter is active, we can't filter prod cats (they're in raw_df by product row)
# So prod categories always show full picture for the date range
prod_sorted   = sorted(prod_cats_raw.items(), key=lambda x: -x[1]["total"])
cat_total_sum = sum(v["total"] for v in prod_cats_raw.values()) if prod_cats_raw else 1

# Transactions metrics
tx_total = tx_received = tx_due = 0.0
tx_payments = {}
if not tx_df.empty:
    tx_total     = float(tx_df["Total Amount"].sum())         if "Total Amount"         in tx_df.columns else 0.0
    tx_received  = float(tx_df["Received/Paid Amount"].sum()) if "Received/Paid Amount" in tx_df.columns else 0.0
    tx_due       = float(tx_df["Balance Due"].sum())          if "Balance Due"          in tx_df.columns else 0.0
    if "Payment Type" in tx_df.columns:
        tx_payments = tx_df.groupby("Payment Type")["Total Amount"].sum().sort_values(ascending=False).to_dict()

collection_rate = (tx_received / tx_total * 100) if tx_total else 0.0

# Period label
if has_real_dates:
    period_label = f"{col_dates[lo-1].strftime('%d %b')} – {col_dates[hi-1].strftime('%d %b %Y')}"
else:
    period_label = f"Day {lo}–{hi}"

# x-axis labels
if has_real_dates:
    x_labels = [col_dates[lo - 1 + i].strftime("%d %b %Y") for i in range(len(daily_totals))]
else:
    x_labels = [str(c) for c in selected_date_cols]

best_date_label  = x_labels[best_idx]  if x_labels and best_idx  < len(x_labels) else f"Day {best_idx+1}"
worst_date_label = x_labels[worst_idx] if x_labels and worst_idx < len(x_labels) else f"Day {worst_idx+1}"

# Active filter summary for display
filter_active = len(names_final) < len(all_pt) if not all_pt.empty else False
filter_note   = f"Showing {len(names_final)} of {len(all_pt)} sellers" if filter_active else f"All {len(all_pt)} sellers"

# MTD progress in sidebar (placed after computation)
with st.sidebar:
    st.markdown(f"""
    <div class="sb-sec" style="margin-top:.3rem;">
      <div class="sb-sec-label">MTD Progress</div><div class="sb-sec-line"></div>
    </div>
    <div class="sb-prog-wrap">
      <div class="sb-prog-track">
        <div class="sb-prog-fill" style="width:{progress_pct:.1f}%;"></div>
      </div>
      <div class="sb-prog-cap">
        <span>₹{real_mtd:,.0f}</span><span>{progress_pct:.1f}%</span>
      </div>
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
# AI INSIGHT GENERATOR
# ─────────────────────────────────────────────────────────────────
def generate_ai_insight(summary_text: str):
    api_key = st.session_state.get("groq_api_key") or os.environ.get("GROQ_API_KEY", "")
    if not api_key:
        return None
    try:
        resp = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "model": "llama-3.3-70b-versatile",
                "messages": [
                    {"role": "system", "content": (
                        "You are ARIA, a sales analytics AI. Given data, write exactly 4 short bullet points. "
                        "Each starts with one emoji. Cover: 1) overall performance 2) standout person or channel "
                        "3) one risk or concern 4) one actionable suggestion. Be specific with numbers. No headers. No markdown."
                    )},
                    {"role": "user", "content": summary_text[:6000]},
                ],
                "temperature": 0.4,
                "max_tokens": 350,
            },
            timeout=20,
        )
        if resp.status_code == 200:
            return resp.json()["choices"][0]["message"]["content"].strip()
        return None
    except Exception:
        return None

@st.cache_data(show_spinner=False, ttl=600)
def cached_ai_insight(cache_key: str, summary_text: str):
    return generate_ai_insight(summary_text)


# ═══════════════════════════════════════════════════════════════════
# ① HERO BANNER
# ═══════════════════════════════════════════════════════════════════
st.markdown(f"""
<div class="hero">
  <div class="hero-inner">
    <div class="hero-status"><div class="hero-sdot"></div>Live Analysis · {period_label} · {hi-lo+1} days · {filter_note}</div>
    <div class="hero-title">Your business,<br><span>decoded in seconds.</span></div>
    <div class="hero-desc">
      ARIA reads your daily report and surfaces what matters — who's leading, where money comes from,
      what's at risk, and what to do next.
    </div>
    <div class="hero-chips">
      <div class="hero-chip">📄 <b>{filename}</b></div>
      <div class="hero-chip">📊 <b>{len(names_final):,}</b> sellers shown</div>
      <div class="hero-chip">👥 <b>{n_active}</b> active</div>
      <div class="hero-chip">🧾 <b>{len(tx_df) if not tx_df.empty else 0}</b> transactions</div>
      <div class="hero-chip">📅 <b>{hi-lo+1}</b> days selected</div>
    </div>
  </div>
  <div class="hero-health {health_cls}">{health_txt}</div>
</div>
""", unsafe_allow_html=True)

# Filter status banner
if filter_active:
    st.markdown(f"""
    <div style="background:rgba(59,130,246,.06);border:1px solid rgba(59,130,246,.18);border-radius:12px;
      padding:.6rem 1.1rem;margin-bottom:1rem;font-size:.76rem;color:#60A5FA;display:flex;align-items:center;gap:8px;">
      🔍 <b>Filter active:</b>&nbsp; {filter_note} · {len(sel_cats)} channel(s) selected.
      All charts below reflect this selection.
    </div>
    """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
# ② AI INSIGHTS
# ═══════════════════════════════════════════════════════════════════
section("AI Insights",
        "ARIA reads your data and gives you a plain-English summary — read this first.",
        anchor_id="sec-ai")

ai_cache_key = f"{filename}-{lo}-{hi}-{total_sales:.0f}-{','.join(sorted(names_final))}"
with st.spinner("ARIA is analysing your data…"):
    ai_text = cached_ai_insight(ai_cache_key, get_data_summary_for_ai(filtered_df))

st.markdown('<div class="ai-card">', unsafe_allow_html=True)
st.markdown("""<div class="ai-card-header">
  <span style="font-size:1.3rem;">🧠</span>
  <div class="ai-card-title">What ARIA sees right now</div>
  <div class="ai-badge">LLaMA 3.3-70B</div>
</div>""", unsafe_allow_html=True)

if ai_text:
    st.markdown(f'<div class="ai-body">{ai_text}</div>', unsafe_allow_html=True)
else:
    fallback = (
        f"📈  Total sales: ₹{total_sales:,.0f}, averaging ₹{avg_daily:,.0f}/day.\n"
        f"🏆  {tp_name} leads the team with ₹{tp_val:,.0f}.\n"
        f"⚠️  {len(zero_list)} salesperson(s) had at least one zero-sales day.\n"
        f"💡  Add your GROQ API key in .env to enable full AI analysis."
    )
    st.markdown(f'<div class="ai-body">{fallback}</div>', unsafe_allow_html=True)
st.markdown("</div>", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
# ③ KEY METRICS
# ═══════════════════════════════════════════════════════════════════
section("Key Metrics",
        "Eight numbers that give you the full picture at a glance.",
        anchor_id="sec-metrics")

c1, c2, c3, c4 = st.columns(4)
with c1:
    st.markdown(f"""<div class="kpi blue">
      <div class="kpi-icon-row"><div class="kpi-icon">💰</div><div class="kpi-hint" title="Total revenue for selected period and filters.">?</div></div>
      <div class="kpi-label">Total Sales</div>
      <div class="kpi-value blue">₹{total_sales:,.0f}</div>
      <div class="kpi-sub">{period_label}</div>
      <div class="kpi-delta {w_cls}">{w_arrow} {abs(week_chg_pct):.1f}% vs last week</div>
    </div>""", unsafe_allow_html=True)
with c2:
    st.markdown(f"""<div class="kpi cyan">
      <div class="kpi-icon-row"><div class="kpi-icon">📅</div><div class="kpi-hint" title="Month-to-date sales vs target.">?</div></div>
      <div class="kpi-label">Month To Date</div>
      <div class="kpi-value cyan">₹{real_mtd:,.0f}</div>
      <div class="kpi-sub">{progress_pct:.1f}% of ₹{monthly_target:,.0f} target</div>
    </div>""", unsafe_allow_html=True)
with c3:
    st.markdown(f"""<div class="kpi gold">
      <div class="kpi-icon-row"><div class="kpi-icon">📊</div><div class="kpi-hint" title="Cumulative sales since Jan 1.">?</div></div>
      <div class="kpi-label">Year To Date</div>
      <div class="kpi-value gold">₹{real_ytd:,.0f}</div>
      <div class="kpi-sub">Cumulative YTD</div>
    </div>""", unsafe_allow_html=True)
with c4:
    st.markdown(f"""<div class="kpi purple">
      <div class="kpi-icon-row"><div class="kpi-icon">⚡</div><div class="kpi-hint" title="Average daily revenue for filtered selection.">?</div></div>
      <div class="kpi-label">Daily Average</div>
      <div class="kpi-value purple">₹{avg_daily:,.0f}</div>
      <div class="kpi-sub">Over {hi-lo+1} days</div>
    </div>""", unsafe_allow_html=True)

st.markdown("<div style='height:.75rem;'></div>", unsafe_allow_html=True)

c5, c6, c7, c8 = st.columns(4)
with c5:
    st.markdown(f"""<div class="kpi gold">
      <div class="kpi-icon-row"><div class="kpi-icon">🏆</div><div class="kpi-hint" title="Highest-selling person in filtered selection.">?</div></div>
      <div class="kpi-label">Top Performer</div>
      <div class="kpi-value gold" style="font-size:1.2rem;">{tp_name}</div>
      <div class="kpi-sub">₹{tp_val:,.0f} total</div>
    </div>""", unsafe_allow_html=True)
with c6:
    st.markdown(f"""<div class="kpi cyan">
      <div class="kpi-icon-row"><div class="kpi-icon">🧾</div><div class="kpi-hint" title="Total invoices generated.">?</div></div>
      <div class="kpi-label">Total Bills</div>
      <div class="kpi-value cyan">{bill_count}</div>
      <div class="kpi-sub">Invoices generated</div>
    </div>""", unsafe_allow_html=True)
with c7:
    pend_color = "red" if tx_due > tx_received else "gold"
    st.markdown(f"""<div class="kpi {pend_color}">
      <div class="kpi-icon-row"><div class="kpi-icon">⏳</div><div class="kpi-hint" title="Money not yet collected.">?</div></div>
      <div class="kpi-label">Outstanding Dues</div>
      <div class="kpi-value {pend_color}">₹{tx_due:,.0f}</div>
      <div class="kpi-sub">{100-collection_rate:.1f}% uncollected</div>
    </div>""", unsafe_allow_html=True)
with c8:
    z_color = "red" if len(zero_list) > 0 else "green"
    st.markdown(f"""<div class="kpi {z_color}">
      <div class="kpi-icon-row"><div class="kpi-icon">🚨</div><div class="kpi-hint" title="Sellers with at least one zero-sales day.">?</div></div>
      <div class="kpi-label">Zero-Sale Alerts</div>
      <div class="kpi-value {z_color}">{len(zero_list)}</div>
      <div class="kpi-sub">Sellers with gaps</div>
    </div>""", unsafe_allow_html=True)

if avg_daily > 0:
    insight(
        f"Daily average ₹{avg_daily:,.0f} → need ₹{max(0, monthly_target - real_mtd):,.0f} more to hit target "
        f"(≈ {max(0, monthly_target - real_mtd) / avg_daily:.0f} more days at current pace).",
        "blue",
    )
else:
    insight("No sales data for selected filters. Try selecting more channels or salespeople.", "gold")


# ═══════════════════════════════════════════════════════════════════
# ④ DAILY SALES TREND
# ═══════════════════════════════════════════════════════════════════
section("Daily Sales Trend",
        "Each bar or point is one day's total revenue for your current filter selection.",
        anchor_id="sec-trend")

if len(daily_totals) and daily_totals.sum() > 0:
    ctrl_left, chart_right = st.columns([1, 4])
    with ctrl_left:
        trend_type = st.radio("Chart Style", ["Area", "Bar", "Line", "Step"], key="trend_type")
        show_ma    = st.checkbox("7-Day Average", value=True)
        show_tgt   = st.checkbox("Daily Target Line", value=False)
        st.markdown(f"""
        <div style="margin-top:1rem;background:rgba(59,130,246,.05);border:1px solid rgba(59,130,246,.14);
          border-radius:12px;padding:.75rem 1rem;font-size:.76rem;line-height:1.85;">
          🟢 <b style="color:var(--txt1);">Best</b><br>{best_date_label}<br>₹{daily_totals.values[best_idx]:,.0f}<br><br>
          🔴 <b style="color:var(--txt1);">Worst</b><br>{worst_date_label}<br>₹{daily_totals.values[worst_idx]:,.0f}
        </div>""", unsafe_allow_html=True)

    with chart_right:
        y_vals       = daily_totals.values
        daily_target = monthly_target / 26 if monthly_target else 0
        fig = go.Figure()
        if trend_type == "Area":
            fig.add_trace(go.Scatter(x=x_labels, y=y_vals, mode="lines",
                line=dict(color="#3B82F6", width=2.5, shape="spline"),
                fill="tozeroy", fillcolor="rgba(59,130,246,.10)", name="Daily Sales"))
        elif trend_type == "Bar":
            colors_b = ["rgba(59,130,246,.85)" if v >= avg_daily else "rgba(139,92,246,.65)" for v in y_vals]
            fig.add_trace(go.Bar(x=x_labels, y=y_vals,
                marker=dict(color=colors_b, line=dict(color="rgba(59,130,246,.3)", width=.5)), name="Daily Sales"))
        elif trend_type == "Line":
            fig.add_trace(go.Scatter(x=x_labels, y=y_vals, mode="lines+markers",
                line=dict(color="#06B6D4", width=2.5, shape="spline"),
                marker=dict(size=5, color="#3B82F6"), name="Daily Sales"))
        else:
            fig.add_trace(go.Scatter(x=x_labels, y=y_vals, mode="lines",
                line=dict(color="#8B5CF6", width=2.5, shape="hv"), name="Daily Sales"))

        if show_ma and len(y_vals) >= 7:
            ma = pd.Series(y_vals).rolling(7).mean().values
            fig.add_trace(go.Scatter(x=x_labels, y=ma, mode="lines",
                line=dict(color="#F59E0B", width=1.5, dash="dot"), name="7-Day Avg"))
        if show_tgt and daily_target:
            fig.add_hline(y=daily_target, line=dict(color="#EF4444", width=1.2, dash="dash"),
                annotation_text=f"Daily Target ₹{daily_target:,.0f}",
                annotation_font=dict(color="#EF4444", size=10))

        fig.update_xaxes(showgrid=False, tickfont=dict(size=9))
        fig.update_yaxes(showgrid=True, gridcolor="rgba(255,255,255,.04)", tickprefix="₹", tickformat=",.0f", tickfont=dict(size=9))
        style_fig(fig, height=340, showlegend=True, legend=dict(orientation="h", y=1.12, x=0, font=dict(size=9.5)))
        st.plotly_chart(fig, use_container_width=True)

    trend_sentence = (
        f"Sales are <b style='color:#10B981;'>growing</b> — up {abs(week_chg_pct):.1f}% vs last week."
        if week_chg_pct >= 0 else
        f"Sales are <b style='color:#EF4444;'>declining</b> — down {abs(week_chg_pct):.1f}% vs last week."
    )
    st.markdown(f"""
    <div class="ai-explain">
      <div class="ai-explain-header"><span>📈</span> ARIA · Daily Sales Trend Analysis</div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:.8rem;margin-bottom:.8rem;">
        <div style="background:rgba(16,185,129,.05);border:1px solid rgba(16,185,129,.15);border-radius:12px;padding:.8rem 1rem;">
          <div style="font-family:'JetBrains Mono',monospace;font-size:.5rem;color:#10B981;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:.4rem;">🟢 Best Day</div>
          <div style="font-size:1.1rem;font-weight:700;color:#10B981;">₹{daily_totals.values[best_idx]:,.0f}</div>
          <div style="font-size:.7rem;color:#475569;">{best_date_label}</div>
        </div>
        <div style="background:rgba(239,68,68,.05);border:1px solid rgba(239,68,68,.15);border-radius:12px;padding:.8rem 1rem;">
          <div style="font-family:'JetBrains Mono',monospace;font-size:.5rem;color:#EF4444;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:.4rem;">🔴 Worst Day</div>
          <div style="font-size:1.1rem;font-weight:700;color:#EF4444;">₹{daily_totals.values[worst_idx]:,.0f}</div>
          <div style="font-size:.7rem;color:#475569;">{worst_date_label}</div>
        </div>
      </div>
      <div class="ai-bullet"><span class="ai-bullet-icon">📊</span><span class="ai-bullet-txt">{trend_sentence}</span></div>
      <div class="ai-bullet"><span class="ai-bullet-icon">📅</span><span class="ai-bullet-txt">Daily average is <b>₹{avg_daily:,.0f}</b>. Days above this are your strong days — look at what happened and repeat it.</span></div>
      <div class="ai-bullet"><span class="ai-bullet-icon">⚡</span><span class="ai-bullet-txt">Gap between best and worst: <b>₹{daily_totals.values[best_idx]-daily_totals.values[worst_idx]:,.0f}</b>. Smaller gap = more consistent business.</span></div>
    </div>
    """, unsafe_allow_html=True)
else:
    st.markdown("""
    <div style="background:rgba(245,158,11,.06);border:1px solid rgba(245,158,11,.15);border-radius:14px;
      padding:1.2rem 1.4rem;font-size:.84rem;color:#F59E0B;">
      ⚠️ No sales data for the current filter selection. Try selecting more channels or salespeople.
    </div>
    """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
# ⑤ SALES CHANNELS + TEAM
# ═══════════════════════════════════════════════════════════════════
section("Sales Channels & Team",
        "Where your revenue comes from and who is generating it — filtered to your current selection.",
        anchor_id="sec-channels")

tab_ch, tab_team = st.tabs(["📡  Sales Channels", "🏆  Team Leaderboard"])

CHANNEL_LABELS = {
    "Office Sales Kerala":     ("🏢", "In-house team · Kerala"),
    "Office Sales Interstate": ("🚚", "In-house · Outside Kerala"),
    "Company Direct Sales":    ("🤝", "Corporate / B2B deals"),
    "ADS":                     ("📣", "Advertising-driven sales"),
    "Exhibition Sales":        ("🎪", "Trade fairs & exhibitions"),
    "E-commerce Sales":        ("🛒", "Amazon · Flipkart · Web"),
    "Medical Store / Shops":   ("💊", "Third-party medical stores"),
}

with tab_ch:
    ch1, ch2 = st.columns([1.4, 1])
    with ch1:
        gc_open("var(--blue)", "Revenue by Channel")
        if not cat_totals.empty:
            cat_total_sum_ch = cat_totals["Total"].sum()
            for _, row in cat_totals.sort_values("Total", ascending=False).iterrows():
                cat_name   = row["Category"]
                pct        = (row["Total"] / cat_total_sum_ch * 100) if cat_total_sum_ch else 0
                icon, desc = CHANNEL_LABELS.get(cat_name, ("📦", "Other"))
                color      = BLUE_PAL[hash(cat_name) % len(BLUE_PAL)]
                st.markdown(f"""
                <div class="ch-row">
                  <div class="ch-name-block"><span class="ch-name">{icon} {cat_name}</span><span class="ch-desc">{desc}</span></div>
                  <div class="ch-bar-wrap"><div class="ch-bar" style="width:{pct:.1f}%;background:{color};"></div></div>
                  <div class="ch-val-block"><div class="ch-val">₹{row['Total']:,.0f}</div><div class="ch-pct">{pct:.1f}%</div></div>
                </div>""", unsafe_allow_html=True)
        else:
            st.markdown("<div style='color:var(--txt3);font-size:.82rem;'>No channel data for current selection.</div>", unsafe_allow_html=True)
        gc_close("The longest bar is your highest-revenue channel.")

    with ch2:
        gc_open("var(--purple)", "Channel Share")
        if not cat_totals.empty:
            donut = px.pie(cat_totals, names="Category", values="Total", hole=0.6,
                color_discrete_sequence=BLUE_PAL)
            donut.update_traces(textfont_color="#F1F5F9", marker=dict(line=dict(color="#020510", width=2)))
            style_fig(donut, height=290, legend=dict(font=dict(size=8.5)))
            st.plotly_chart(donut, use_container_width=True)
        gc_close()

    if not cat_totals.empty:
        top_ch     = cat_totals.sort_values("Total", ascending=False).iloc[0]
        bottom_ch  = cat_totals.sort_values("Total", ascending=False).iloc[-1]
        top_ch_pct = top_ch["Total"] / cat_totals["Total"].sum() * 100
        conc_msg   = 'Revenue is <b>heavily concentrated</b> — consider diversifying.' if top_ch_pct > 50 else 'Revenue is <b>well spread</b> across channels — healthy diversification.'
        st.markdown(f"""
        <div class="ai-explain">
          <div class="ai-explain-header"><span>📡</span> ARIA · Channel Intelligence</div>
          <div class="ai-bullet"><span class="ai-bullet-icon">🏆</span>
            <span class="ai-bullet-txt"><b>{top_ch["Category"]}</b> is your #1 channel at ₹{top_ch["Total"]:,.0f} ({top_ch_pct:.1f}% of total).</span></div>
          <div class="ai-bullet"><span class="ai-bullet-icon">⚖️</span>
            <span class="ai-bullet-txt">{conc_msg}</span></div>
          <div class="ai-bullet"><span class="ai-bullet-icon">📉</span>
            <span class="ai-bullet-txt">Lowest channel: <b>{bottom_ch["Category"]}</b> at ₹{bottom_ch["Total"]:,.0f}. Investigate why.</span></div>
        </div>
        """, unsafe_allow_html=True)

# Team section — anchor placed OUTSIDE tab so navigation works
with tab_team:
    # Anchor for team nav link
    st.markdown('<span id="sec-team"></span>', unsafe_allow_html=True)
    t1, t2 = st.columns([1.6, 1])
    with t1:
        gc_open("var(--blue)", "Salesperson Leaderboard — Filtered Selection")
        if not person_totals.empty:
            max_v  = person_totals["Total"].max()
            medals = {0: "r1", 1: "r2", 2: "r3"}
            html   = ""
            for i, row in person_totals.reset_index(drop=True).iterrows():
                pct      = (row["Total"] / max_v * 100) if max_v else 0
                rank_cls = medals.get(i, "")
                av_cls   = "top" if i == 0 else ""
                init     = row["Salesperson"][0].upper() if row["Salesperson"] else "?"
                watch    = row["Salesperson"] in zero_list
                tag      = '<span class="tag-watch">⚠ Watch</span>' if watch else '<span class="tag-ok">✓ Reliable</span>'
                html += f"""
                <div class="lb-row">
                  <div class="lb-rank {rank_cls}">#{i+1}</div>
                  <div class="lb-av {av_cls}">{init}</div>
                  <div class="lb-info">
                    <div class="lb-row-top">
                      <div><span class="lb-name">{row['Salesperson']}</span>
                           <span class="lb-cat"> · {row['Category']}</span>{tag}</div>
                      <span class="lb-amount">₹{row['Total']:,.0f}</span>
                    </div>
                    <div class="lb-bar-track"><div class="lb-bar-fill" style="width:{pct:.1f}%;"></div></div>
                  </div>
                </div>"""
            st.markdown(html, unsafe_allow_html=True)
        else:
            st.markdown("<div style='color:var(--txt3);font-size:.82rem;'>No sellers match the current filter.</div>", unsafe_allow_html=True)
        gc_close("'Watch' = at least one zero-sales day in the selected period.")

        if not person_totals.empty:
            top_p          = person_totals.iloc[0]
            top_p_share    = top_p["Total"] / person_totals["Total"].sum() * 100 if person_totals["Total"].sum() else 0
            watch_count    = len(zero_list)
            avg_per_person = float(person_totals["Total"].mean())
            watch_alert    = (
                f'<b style="color:#EF4444;">{watch_count} seller(s)</b> have zero-sale days — follow up immediately.'
                if watch_count > 0 else
                '✅ <b>No zero-sale alerts</b> — full team is active!'
            )
            st.markdown(f"""
            <div class="ai-explain" style="margin-top:1rem;">
              <div class="ai-explain-header"><span>🏆</span> ARIA · Team Performance</div>
              <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:.6rem;margin-bottom:.8rem;">
                <div style="background:rgba(245,158,11,.05);border:1px solid rgba(245,158,11,.15);border-radius:11px;padding:.7rem .8rem;text-align:center;">
                  <div style="font-size:.55rem;color:#F59E0B;font-family:'JetBrains Mono',monospace;text-transform:uppercase;letter-spacing:1px;margin-bottom:.3rem;">⭐ Leader</div>
                  <div style="font-size:.78rem;font-weight:700;color:#F59E0B;">{top_p['Salesperson'][:14]}</div>
                </div>
                <div style="background:rgba(59,130,246,.05);border:1px solid rgba(59,130,246,.15);border-radius:11px;padding:.7rem .8rem;text-align:center;">
                  <div style="font-size:.55rem;color:#60A5FA;font-family:'JetBrains Mono',monospace;text-transform:uppercase;letter-spacing:1px;margin-bottom:.3rem;">Avg/Person</div>
                  <div style="font-size:.85rem;font-weight:700;color:#60A5FA;">₹{avg_per_person:,.0f}</div>
                </div>
                <div style="background:rgba(239,68,68,.05);border:1px solid rgba(239,68,68,.15);border-radius:11px;padding:.7rem .8rem;text-align:center;">
                  <div style="font-size:.55rem;color:#EF4444;font-family:'JetBrains Mono',monospace;text-transform:uppercase;letter-spacing:1px;margin-bottom:.3rem;">⚠ Watch</div>
                  <div style="font-size:1.3rem;font-weight:700;color:#EF4444;">{watch_count}</div>
                </div>
              </div>
              <div class="ai-bullet"><span class="ai-bullet-icon">🥇</span>
                <span class="ai-bullet-txt"><b>{top_p['Salesperson']}</b> leads with ₹{top_p['Total']:,.0f} — {top_p_share:.1f}% of filtered team sales.</span></div>
              <div class="ai-bullet"><span class="ai-bullet-icon">📊</span>
                <span class="ai-bullet-txt">Team average: <b>₹{avg_per_person:,.0f}/person</b>. Anyone below this needs a check-in.</span></div>
              <div class="ai-bullet"><span class="ai-bullet-icon">⚠️</span>
                <span class="ai-bullet-txt">{watch_alert}</span></div>
            </div>
            """, unsafe_allow_html=True)

    with t2:
        gc_open("var(--red)", "⚠️ Zero-Sale Alerts")
        if zero_list:
            for n in zero_list[:8]:
                st.markdown(f'<div class="alert-row"><div class="alert-dot"></div>{n} · zero-sale day detected</div>', unsafe_allow_html=True)
            if len(zero_list) > 8:
                st.markdown(f"<div style='font-size:.7rem;color:var(--txt3);margin-top:.4rem;'>+{len(zero_list)-8} more sellers</div>", unsafe_allow_html=True)
        else:
            st.markdown('<div style="color:var(--green);font-size:.82rem;">✅ No zero-sales days in selected period</div>', unsafe_allow_html=True)
        gc_close()

        st.markdown("<div style='height:.6rem;'></div>", unsafe_allow_html=True)
        st.markdown(f"""
        <div class="gc">
          <div class="gc-title"><div class="gc-dot" style="background:var(--gold);"></div>🎯 Monthly Target</div>
          <div class="prog-track"><div class="prog-fill" style="width:{progress_pct:.1f}%;"></div></div>
          <div class="prog-cap"><span>₹{real_mtd:,.0f} achieved</span><span>{progress_pct:.1f}%</span></div>
          <div style="margin-top:.65rem;font-size:.75rem;color:var(--txt3);">
            Remaining: <b style="color:var(--txt1);">₹{max(0, monthly_target-real_mtd):,.0f}</b>
          </div>
        </div>""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
# ⑥ PRODUCT CATEGORIES
# ═══════════════════════════════════════════════════════════════════
# Anchor placed here so it's visible outside tabs
st.markdown('<span id="sec-products"></span>', unsafe_allow_html=True)
section("Product Category Analysis",
        "Which product categories are driving your revenue — ranked from highest to lowest.",
        anchor_id="")

pc1, pc2 = st.columns([1.5, 1])

with pc1:
    st.markdown('<div class="gc"><div class="gc-title"><div class="gc-dot" style="background:var(--blue);"></div>Revenue by Product Category — Ranked</div>', unsafe_allow_html=True)
    if prod_sorted:
        max_prod_val = prod_sorted[0][1]["total"]
        for i, (key, meta) in enumerate(prod_sorted):
            pct    = meta["total"] / cat_total_sum * 100 if cat_total_sum else 0
            bar_w  = meta["total"] / max_prod_val * 100
            is_top = i == 0

            safe_color = meta.get("color") or "#3B82F6"
            if not (isinstance(safe_color, str) and safe_color.startswith("#") and len(safe_color) in (4, 7)):
                safe_color = "#3B82F6"
            safe_emoji = meta.get("emoji") or "📦"
            safe_label = meta.get("label") or str(key)
            safe_total = meta.get("total") or 0
            safe_bar_w = max(0.0, min(100.0, bar_w))

            top_badge = (
                '<span style="font-size:.6rem;background:rgba(245,158,11,.1);color:#F59E0B;'
                'border:1px solid rgba(245,158,11,.2);border-radius:6px;'
                'padding:1px 7px;margin-left:4px;">⭐ TOP</span>'
                if is_top else ""
            )
            card_html = (
                f'<div class="cat-bar-card" style="border-left:3px solid {safe_color}40;">'
                f'<div class="cat-bar-card-top">'
                f'<div class="cat-bar-name">'
                f'<span style="font-size:1.1rem;">{safe_emoji}</span> {safe_label}{top_badge}'
                f'</div>'
                f'<div class="cat-bar-right">'
                f'<div class="cat-bar-val" style="color:{safe_color};">₹{safe_total:,.0f}</div>'
                f'<div class="cat-bar-pct">{pct:.1f}% of total</div>'
                f'</div>'
                f'</div>'
                f'<div class="cat-bar-track">'
                f'<div class="cat-bar-fill" style="width:{safe_bar_w:.1f}%;'
                f'background:linear-gradient(90deg,{safe_color},{safe_color}88);"></div>'
                f'</div>'
                f'</div>'
            )
            st.markdown(card_html, unsafe_allow_html=True)
    else:
        st.markdown("<div style='color:var(--txt3);font-size:.82rem;'>No product category data found in this file.</div>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

with pc2:
    st.markdown('<div class="gc"><div class="gc-title"><div class="gc-dot" style="background:var(--purple);"></div>Share Breakdown</div>', unsafe_allow_html=True)
    if prod_sorted:
        labels_p = [v["label"] for _, v in prod_sorted]
        values_p = [v["total"] for _, v in prod_sorted]
        colors_p = [v["color"] for _, v in prod_sorted]
        donut_p  = go.Figure(go.Pie(
            labels=labels_p, values=values_p, hole=0.62,
            marker=dict(colors=colors_p, line=dict(color="#020510", width=2)),
            textfont=dict(size=9, color="#F1F5F9"),
        ))
        donut_p.update_traces(textposition="outside", texttemplate="%{label}<br>%{percent}")
        n_cats_p = len(prod_sorted)
        style_fig(donut_p, height=260, showlegend=False,
            annotations=[dict(
                text=f"<b>{n_cats_p}</b><br>Categories",
                x=0.5, y=0.5, font_size=11, showarrow=False, font_color="#94A3B8",
            )])
        st.plotly_chart(donut_p, use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)

    if prod_sorted:
        top_cat     = prod_sorted[0]
        top_cat_pct = top_cat[1]["total"] / cat_total_sum * 100 if cat_total_sum else 0
        bottom_cat  = prod_sorted[-1]
        cumsum      = 0
        vital_few   = []
        for key, meta in prod_sorted:
            cumsum += meta["total"]
            vital_few.append(meta["label"])
            if cat_total_sum and cumsum / cat_total_sum >= 0.8:
                break
        vital_word  = "category" if len(vital_few) == 1 else "categories"
        vital_names = " · ".join(vital_few)
        st.markdown(f"""
        <div class="ai-explain" style="margin-top:.8rem;">
          <div class="ai-explain-header"><span>🧬</span> ARIA · Product Insights</div>
          <div class="ai-bullet">
            <span class="ai-bullet-icon">⭐</span>
            <span class="ai-bullet-txt"><b>{top_cat[1]['label']}</b> is #1 at ₹{top_cat[1]['total']:,.0f} ({top_cat_pct:.1f}% share).</span>
          </div>
          <div class="ai-bullet">
            <span class="ai-bullet-icon">📌</span>
            <span class="ai-bullet-txt"><b>{len(vital_few)} {vital_word}</b> generate 80% of product revenue: {vital_names}</span>
          </div>
          <div class="ai-bullet">
            <span class="ai-bullet-icon">💡</span>
            <span class="ai-bullet-txt">Lowest: <b>{bottom_cat[1]['label']}</b> at ₹{bottom_cat[1]['total']:,.0f}. Invest or shift resources.</span>
          </div>
        </div>
        """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
# ⑦ RETURNS & NET REVENUE
# ═══════════════════════════════════════════════════════════════════
section("Returns & Net Revenue",
        "How much of your gross sales actually stayed as real revenue after returns.",
        anchor_id="sec-returns")

return_pct   = (returns_total / total_sales * 100) if total_sales else 0
retained_pct = 100 - return_pct
r_color      = "#EF4444" if return_pct > 10 else ("#F59E0B" if return_pct > 5 else "#10B981")
r_msg        = "⚠️ High — Investigate immediately" if return_pct > 10 else ("🟡 Moderate — Monitor closely" if return_pct > 5 else "✅ Healthy — Keep it up")

st.markdown(f"""
<div class="flow-steps">
  <div class="flow-step gross"><div class="flow-step-icon">💰</div>
    <div class="flow-step-label gross">Gross Sales</div>
    <div class="flow-step-val gross">₹{total_sales:,.0f}</div>
    <div class="flow-step-sub">Total invoiced</div></div>
  <div class="flow-arrow">→</div>
  <div class="flow-step ret"><div class="flow-step-icon">↩️</div>
    <div class="flow-step-label ret">Returns</div>
    <div class="flow-step-val ret">−₹{returns_total:,.0f}</div>
    <div class="flow-step-sub">{return_pct:.1f}% lost</div></div>
  <div class="flow-arrow">→</div>
  <div class="flow-step net"><div class="flow-step-icon">✅</div>
    <div class="flow-step-label net">Net Revenue</div>
    <div class="flow-step-val net">₹{net_sales:,.0f}</div>
    <div class="flow-step-sub">{retained_pct:.1f}% retained</div></div>
</div>
""", unsafe_allow_html=True)

rn1, rn2 = st.columns([1.6, 1])
with rn1:
    gc_open("var(--red)", "Revenue Flow — Gross to Net")
    wf = go.Figure(go.Waterfall(
        orientation="v", measure=["absolute","relative","total"],
        x=["💰 Gross Sales","↩️ Returns","✅ Net Revenue"],
        y=[total_sales, -returns_total, net_sales],
        connector=dict(line=dict(color="rgba(255,255,255,0.1)", width=1)),
        decreasing=dict(marker=dict(color="#EF4444")),
        increasing=dict(marker=dict(color="#10B981")),
        totals=dict(marker=dict(color="#3B82F6")),
        text=[f"₹{total_sales:,.0f}", f"−₹{returns_total:,.0f}", f"₹{net_sales:,.0f}"],
        textposition="outside", textfont=dict(color="#F1F5F9", size=11),
    ))
    wf.update_yaxes(tickprefix="₹", tickformat=",.0f", showgrid=True, gridcolor="rgba(255,255,255,.04)")
    wf.update_xaxes(showgrid=False)
    style_fig(wf, height=320, showlegend=False)
    st.plotly_chart(wf, use_container_width=True)
    st.markdown(f"""
    <div style="margin-top:.5rem;padding:.7rem 1rem;border-radius:12px;background:rgba(16,185,129,.04);border:1px solid rgba(16,185,129,.14);">
      <div style="display:flex;justify-content:space-between;margin-bottom:.4rem;">
        <span style="font-size:.75rem;color:#94A3B8;">Revenue retained after returns</span>
        <span style="font-family:'JetBrains Mono',monospace;font-size:.75rem;color:#10B981;font-weight:700;">{retained_pct:.1f}%</span>
      </div>
      <div style="height:8px;border-radius:99px;background:rgba(255,255,255,.05);overflow:hidden;">
        <div style="height:100%;width:{retained_pct:.1f}%;background:linear-gradient(90deg,#10B981,#06B6D4);border-radius:99px;"></div>
      </div>
    </div>""", unsafe_allow_html=True)
    gc_close()

with rn2:
    st.markdown(f"""
    <div class="gc">
      <div class="gc-title"><div class="gc-dot" style="background:var(--red);"></div>Return Rate</div>
      <div style="font-size:3rem;font-weight:700;color:{r_color};font-family:'Space Grotesk',sans-serif;">{return_pct:.1f}%</div>
      <div style="font-size:.7rem;color:#475569;margin-bottom:.8rem;">of gross sales returned</div>
      <div style="font-size:.75rem;color:{r_color};margin-bottom:.8rem;">{r_msg}</div>
      <div class="coll-row"><span class="coll-lbl">Gross Sales</span><span class="coll-val" style="color:#60A5FA;">₹{total_sales:,.0f}</span></div>
      <div class="coll-row"><span class="coll-lbl">Returns</span><span class="coll-val" style="color:#EF4444;">−₹{returns_total:,.0f}</span></div>
      <div class="coll-row"><span class="coll-lbl">Net Revenue</span><span class="coll-val" style="color:#10B981;">₹{net_sales:,.0f}</span></div>
      <div class="coll-row"><span class="coll-lbl">Retained</span><span class="coll-val" style="color:#10B981;">{retained_pct:.1f}%</span></div>
    </div>""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
# ⑧ TRANSACTIONS
# ═══════════════════════════════════════════════════════════════════
section("Transactions & Collections",
        "Who owes you money, how they're paying, and which follow-ups are due.",
        anchor_id="sec-transactions")

if not tx_df.empty and show_tx:
    tc1, tc2, tc3 = st.columns(3)
    with tc1:
        gc_open("var(--green)", "💳 Collection Status")
        if tx_total > 0:
            coll_color = "var(--green)" if collection_rate >= 70 else "var(--gold)"
            st.markdown(f"""
            <div class="coll-row"><span class="coll-lbl">Total Invoiced</span><span class="coll-val">₹{tx_total:,.0f}</span></div>
            <div class="coll-row"><span class="coll-lbl">Collected</span><span class="coll-val" style="color:var(--green);">₹{tx_received:,.0f}</span></div>
            <div class="coll-row"><span class="coll-lbl">Outstanding</span><span class="coll-val" style="color:var(--red);">₹{tx_due:,.0f}</span></div>
            <div class="coll-row"><span class="coll-lbl">Collection Rate</span>
              <span class="coll-val" style="color:{coll_color};">{collection_rate:.1f}%</span></div>
            """, unsafe_allow_html=True)
            donut2 = go.Figure(go.Pie(
                values=[tx_received, tx_due], labels=["Collected","Outstanding"], hole=0.62,
                marker=dict(colors=["#10B981","#EF4444"], line=dict(color="rgba(0,0,0,0)", width=0)),
                textfont=dict(size=10),
            ))
            style_fig(donut2, height=160, showlegend=False, margin=dict(l=0, r=0, t=5, b=5))
            st.plotly_chart(donut2, use_container_width=True)
            if collection_rate >= 70:
                coll_insight = f'Collection rate is <b style="color:#10B981;">healthy</b> at {collection_rate:.1f}%. Keep following up.'
            else:
                coll_insight = f'Collection rate of <b style="color:#EF4444;">{collection_rate:.1f}%</b> is below 70% benchmark. Prioritise collections.'
            st.markdown(f"""
            <div class="ai-explain" style="margin-top:.8rem;">
              <div class="ai-explain-header"><span>💰</span> ARIA · Collection Health</div>
              <div class="ai-bullet"><span class="ai-bullet-icon">💵</span>
                <span class="ai-bullet-txt">Collected <b style="color:#10B981;">₹{tx_received:,.0f}</b> of ₹{tx_total:,.0f}.
                <b style="color:#EF4444;">₹{tx_due:,.0f}</b> still outstanding.</span></div>
              <div class="ai-bullet"><span class="ai-bullet-icon">📊</span>
                <span class="ai-bullet-txt">{coll_insight}</span></div>
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown("<div style='color:var(--txt3);font-size:.8rem;'>No transaction data.</div>", unsafe_allow_html=True)
        gc_close()

    with tc2:
        gc_open("var(--cyan)", "💰 Payment Methods")
        if tx_payments:
            total_pay  = sum(tx_payments.values())
            PAY_COLORS = {"Razorpay":"#8B5CF6","Cash":"#10B981","Federal Bank":"#3B82F6"}
            for pay, val in tx_payments.items():
                pct   = val / total_pay * 100 if total_pay else 0
                color = PAY_COLORS.get(pay, "#06B6D4")
                st.markdown(f"""
                <div class="ch-row">
                  <div class="ch-name-block"><span class="ch-name">{pay}</span></div>
                  <div class="ch-bar-wrap"><div class="ch-bar" style="width:{pct:.1f}%;background:{color};"></div></div>
                  <div class="ch-val-block"><div class="ch-val">₹{val:,.0f}</div><div class="ch-pct">{pct:.1f}%</div></div>
                </div>""", unsafe_allow_html=True)
        else:
            st.markdown("<div style='color:var(--txt3);font-size:.8rem;'>No payment data.</div>", unsafe_allow_html=True)
        gc_close()

    with tc3:
        gc_open("var(--gold)", "📋 Upcoming Follow-Ups")
        if "Follow up date" in tx_df.columns and "Balance Due" in tx_df.columns:
            fu_df = tx_df[tx_df["Balance Due"] > 0].sort_values("Balance Due", ascending=False)
            for _, frow in fu_df.head(6).iterrows():
                party   = frow.get("Party Name", "—")
                due     = frow.get("Balance Due", 0)
                fu_date = frow.get("Follow up date", "—")
                st.markdown(f"""
                <div class="fu-row"><div class="fu-dot"></div>
                  <div style="flex:1;">
                    <div class="fu-party">{party} · <b style="color:var(--red);">₹{due:,.0f}</b></div>
                    <div class="fu-due">Follow up: {fu_date}</div>
                  </div>
                </div>""", unsafe_allow_html=True)
            st.markdown(f"""
            <div class="ai-explain" style="margin-top:.8rem;">
              <div class="ai-explain-header"><span>📅</span> ARIA · Follow-Up Guide</div>
              <div class="ai-bullet"><span class="ai-bullet-icon">🔔</span>
                <span class="ai-bullet-txt"><b style="color:#F59E0B;">{len(fu_df)} open invoices</b> · ₹{fu_df["Balance Due"].sum():,.0f} outstanding.</span></div>
              <div class="ai-bullet"><span class="ai-bullet-icon">📋</span>
                <span class="ai-bullet-txt">List sorted by <b>highest balance first</b> — call biggest amounts first for max cash recovery.</span></div>
              <div class="ai-bullet"><span class="ai-bullet-icon">⏰</span>
                <span class="ai-bullet-txt">Call <b>before 11am</b> — decision makers most reachable in the morning.</span></div>
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown("<div style='color:var(--txt3);font-size:.8rem;'>No follow-up data.</div>", unsafe_allow_html=True)
        gc_close()
else:
    st.markdown("<div style='color:var(--txt3);font-size:.82rem;padding:1rem;'>No transaction sheet found or transactions hidden.</div>", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
# ⑨ EXPORT & DATA TABLES
# ═══════════════════════════════════════════════════════════════════
section("Export & Data Tables",
        "Download your data or explore the raw numbers directly.",
        anchor_id="sec-export")

e1, e2, e3 = st.columns(3)
with e1:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        filtered_df.to_excel(writer, index=False, sheet_name="Filtered")
        person_totals.to_excel(writer, index=False, sheet_name="Leaderboard")
        cat_totals.to_excel(writer, index=False, sheet_name="Categories")
        if not tx_df.empty:
            tx_df.to_excel(writer, index=False, sheet_name="Transactions")
    st.markdown('<div class="export-card"><div class="export-icon">📊</div><div class="export-title">Full Excel Report</div><div class="export-desc">All sheets: filtered data, leaderboard, categories, transactions.</div>', unsafe_allow_html=True)
    st.download_button("⬇ Download Excel", data=buf.getvalue(),
        file_name=f"ARIA_Report_{filename.split('.')[0]}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)

with e2:
    csv_buf = BytesIO()
    filtered_df.to_csv(csv_buf, index=False)
    st.markdown('<div class="export-card"><div class="export-icon">📄</div><div class="export-title">CSV Export</div><div class="export-desc">Filtered dataset as CSV — open in Excel, Google Sheets, or any tool.</div>', unsafe_allow_html=True)
    st.download_button("⬇ Download CSV", data=csv_buf.getvalue(),
        file_name=f"ARIA_{filename.split('.')[0]}.csv",
        mime="text/csv", use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)

with e3:
    ai_summary = get_data_summary_for_ai(filtered_df)
    st.markdown('<div class="export-card"><div class="export-icon">🧠</div><div class="export-title">AI Summary Text</div><div class="export-desc">Plain-text data summary used by ARIA — paste into any AI tool.</div>', unsafe_allow_html=True)
    st.download_button("⬇ Download AI Summary", data=ai_summary,
        file_name="ARIA_AI_Context.txt", mime="text/plain", use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)

if show_raw:
    st.markdown("<div style='height:1rem;'></div>", unsafe_allow_html=True)
    tbl_view = st.radio(
        "View Table",
        ["Salesperson Totals","Category Totals","Daily Summary","Filtered Dataset","Full Dataset"],
        horizontal=True, key="tbl_v",
    )
    if tbl_view == "Salesperson Totals":
        st.dataframe(person_totals.style.format({"Total":"₹{:,.0f}"}), use_container_width=True, height=380)
    elif tbl_view == "Category Totals":
        st.dataframe(cat_totals.style.format({"Total":"₹{:,.0f}"}), use_container_width=True, height=380)
    elif tbl_view == "Daily Summary":
        daily_df = pd.DataFrame({
            "Date":      x_labels,
            "Sales":     daily_totals.values,
            "7-Day Avg": pd.Series(daily_totals.values).rolling(7).mean().values,
        })
        st.dataframe(daily_df.style.format({"Sales":"₹{:,.0f}","7-Day Avg":"₹{:,.0f}"}),
            use_container_width=True, height=380)
    elif tbl_view == "Filtered Dataset":
        st.dataframe(filtered_df, use_container_width=True, height=380)
    else:
        st.dataframe(df, use_container_width=True, height=380)

if show_tx and not tx_df.empty:
    st.markdown("<div style='height:.8rem;'></div>", unsafe_allow_html=True)
    tf1, tf2, tf3 = st.columns(3)
    with tf1:
        ch_opts = sorted(tx_df["Sales Channel"].dropna().unique().tolist()) if "Sales Channel" in tx_df.columns else []
        ch_fil  = st.multiselect("Channel filter", ch_opts, default=None, placeholder="All channels", label_visibility="collapsed")
    with tf2:
        sm_opts = sorted(tx_df["Salesman Name"].dropna().unique().tolist()) if "Salesman Name" in tx_df.columns else []
        sm_fil  = st.multiselect("Salesman filter", sm_opts, default=None, placeholder="All salesmen", label_visibility="collapsed")
    with tf3:
        pending_only = st.checkbox("Pending only (Balance Due > 0)", value=False)

    disp_tx = tx_df.copy()
    if ch_fil and "Sales Channel" in disp_tx.columns:
        disp_tx = disp_tx[disp_tx["Sales Channel"].isin(ch_fil)]
    if sm_fil and "Salesman Name" in disp_tx.columns:
        disp_tx = disp_tx[disp_tx["Salesman Name"].isin(sm_fil)]
    if pending_only and "Balance Due" in disp_tx.columns:
        disp_tx = disp_tx[disp_tx["Balance Due"] > 0]

    st.dataframe(disp_tx, use_container_width=True, height=400)
    total_amt = disp_tx["Total Amount"].sum() if "Total Amount" in disp_tx.columns else 0
    total_due = disp_tx["Balance Due"].sum()  if "Balance Due"  in disp_tx.columns else 0
    st.markdown(
        f"<div style='font-size:.7rem;color:var(--txt3);margin-top:.3rem;'>"
        f"{len(disp_tx):,} transactions · ₹{total_amt:,.0f} total · ₹{total_due:,.0f} outstanding</div>",
        unsafe_allow_html=True,
    )


# ─────────────────────────────────────────────────────────────────
# FOOTER
# ─────────────────────────────────────────────────────────────────
st.markdown("""
<div class="aria-footer">ARIA · Sales Intelligence Platform · v5.0 Neural Glass · Built with ❤️</div>
""", unsafe_allow_html=True)